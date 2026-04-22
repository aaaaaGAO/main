#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 工作簿底层封装（openpyxl）

将 openpyxl 的打开、合并单元格取值等「硬核」逻辑收拢在 infra 层，
与 utils 中通用字符串处理（norm_str、nfc_normalize 等）分离。
"""

from __future__ import annotations

import io
import os
from typing import Any

from openpyxl import load_workbook
from infra.filesystem import resolve_runtime_path


def merged_cell_value(ws: Any, row: int, col: int) -> Any:
    """取单元格值，支持合并单元格（空值取合并区域左上角值）。
    参数：ws — 工作表；row, col — 行号、列号。
    返回：单元格值。
    """
    cell = ws.cell(row, col)
    cell_value = cell.value
    if cell_value is not None:
        return cell_value
    try:
        coord = cell.coordinate
        for merged_range in getattr(ws.merged_cells, "ranges", []):
            if coord in merged_range:
                return ws.cell(merged_range.min_row, merged_range.min_col).value
    except Exception:
        pass
    return cell_value


class ExcelService:
    """统一的 Excel 工作簿打开与行迭代（openpyxl 封装）。"""
    _WORKBOOK_BINARY_CACHE: dict[tuple[str, float], bytes] = {}

    @staticmethod
    def open_workbook(
        excel_file_path: str,
        *,
        read_only: bool = True,
        data_only: bool = True,
        rich_text: bool = False,
        nfc: bool = False,
    ):
        """打开 Excel 工作簿，统一处理常见异常。
        参数：excel_file_path — 文件路径；read_only — 只读模式；data_only — 只读公式结果；rich_text — 富文本；nfc — 是否 NFC 归一化。
        返回：openpyxl Workbook。异常：FileNotFoundError / ValueError / PermissionError。
        """
        excel_file_path = resolve_runtime_path(None, excel_file_path)

        if not os.path.exists(excel_file_path):
            raise FileNotFoundError(f"找不到 Excel 文件: {excel_file_path}")
        if not os.path.isfile(excel_file_path):
            raise ValueError(f"路径不是文件: {excel_file_path}")
        if not excel_file_path.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
            raise ValueError(f"文件不是有效的 Excel 文件: {excel_file_path}")

        kwargs = {"data_only": data_only, "read_only": read_only}
        if rich_text:
            kwargs["rich_text"] = True
            kwargs["read_only"] = False

        try:
            if read_only and not rich_text:
                path_mtime = os.path.getmtime(excel_file_path)
                cache_key = (excel_file_path, path_mtime)
                payload_data = ExcelService._WORKBOOK_BINARY_CACHE.get(cache_key)
                if payload_data is None:
                    with open(excel_file_path, "rb") as fp:
                        payload_data = fp.read()
                    ExcelService._WORKBOOK_BINARY_CACHE = {cache_key: payload_data}
                wb = load_workbook(io.BytesIO(payload_data), **kwargs)
            else:
                wb = load_workbook(excel_file_path, **kwargs)
            try:
                for sheet_name in wb.sheetnames:
                    try:
                        worksheet = wb[sheet_name]
                        if getattr(worksheet, "_dimensions", None) is not None:
                            worksheet._dimensions = None
                    except Exception:
                        continue
            except Exception:
                pass
            return wb
        except FileNotFoundError:
            raise FileNotFoundError(f"找不到文件: {excel_file_path}")
        except PermissionError:
            raise PermissionError(f"没有权限读取文件: {excel_file_path}\n请确认文件未被其他程序打开")
        except Exception as error:
            error_msg = str(error).lower()
            if any(
                kw in error_msg
                for kw in ("decompressing", "incorrect header", "badzipfile", "not a zip file")
            ):
                raise ValueError(
                    f"Excel 文件格式错误或文件已损坏: {excel_file_path}\n"
                    f"错误详情: {error}\n"
                    f"请检查文件是否是有效的 Excel 文件（.xlsx 格式）"
                )
            raise ValueError(f"无法读取 Excel 文件: {excel_file_path}\n错误详情: {error}")

    @staticmethod
    def iter_rows(
        file_path: str,
        sheet_name: str,
        *,
        min_row: int = 1,
        max_col: int | None = None,
        values_only: bool = True,
    ):
        """打开 Excel 并迭代指定 sheet 的行。
        参数：file_path — 文件路径；sheet_name — Sheet 名；min_row, max_col, values_only — 迭代选项。
        返回：行迭代器。
        """
        wb = ExcelService.open_workbook(file_path)
        ws = wb[sheet_name]
        kwargs = {"min_row": min_row, "values_only": values_only}
        if max_col is not None:
            kwargs["max_col"] = max_col
        return ws.iter_rows(**kwargs)

    @staticmethod
    def resolve_sheets(wb: Any, sheets_raw: str) -> list:
        """解析 sheet 筛选字符串：* 或空为全部，否则为指定列表。
        参数：wb — 工作簿；sheets_raw — 配置字符串。
        返回：Sheet 名列表。
        """
        if not sheets_raw or sheets_raw.strip() in ("*", ""):
            return list(wb.sheetnames)
        sheet_candidates = [
            sheet_text.strip()
            for sheet_text in sheets_raw.split(",")
            if sheet_text.strip()
        ]
        existing_sheets = [
            sheet_name for sheet_name in sheet_candidates if sheet_name in wb.sheetnames
        ]
        return existing_sheets if existing_sheets else list(wb.sheetnames)


__all__ = ["ExcelService", "merged_cell_value"]
