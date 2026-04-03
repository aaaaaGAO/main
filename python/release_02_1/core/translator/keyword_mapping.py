#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
关键字-CAPL 映射读取器。

供 CAN/CIN 等生成器复用，统一读取「关键字-CAPL函数映射表.xlsx」。
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Callable

from openpyxl import load_workbook
from infra.excel.header import normalize_cell_text


@dataclass(slots=True)
class KeywordSpec:
    """一条关键字到 CAPL 函数的映射。"""

    func_name: str
    keyword: str
    capl_func: str
    remark: str = ""
    source_sheet: str = ""

    @property
    def full_key(self) -> str:
        if self.func_name and self.keyword:
            return f"{self.func_name}::{self.keyword}"
        if self.func_name:
            return self.func_name
        if self.keyword:
            return f"::{self.keyword}"
        return ""

def load_keyword_specs_from_excel(
    excel_path: str,
    sheet_names: list[str],
    *,
    warn: Callable[[str], None] | None = None,
) -> dict[str, KeywordSpec]:
    """从「关键字-CAPL函数映射表」Excel 读取多 Sheet，返回 full_key.lower() -> KeywordSpec 的字典。
    参数: excel_path — 映射表 xlsx 路径；sheet_names — 要读取的 Sheet 名列表；warn — 可选告警回调。
    返回: 关键字规格字典。文件不存在或损坏时按 warn 提示并返回空字典或抛 ValueError。
    """
    specs: dict[str, KeywordSpec] = {}
    if not excel_path:
        return specs

    try:
        wb = load_workbook(excel_path, data_only=True)
    except FileNotFoundError:
        if warn:
            warn(f"未找到映射表: {excel_path}")
        return specs
    except Exception as error:
        error_msg = str(error)
        if (
            "decompressing" in error_msg.lower()
            or "incorrect header" in error_msg.lower()
            or "badzipfile" in error_msg.lower()
        ):
            raise ValueError(
                f"映射表 Excel 文件格式错误或文件已损坏: {excel_path}\n"
                f"错误详情: {error_msg}\n"
                "请检查文件是否是有效的 Excel 文件（.xlsx 格式）"
            ) from error
        raise ValueError(
            f"无法读取映射表 Excel 文件: {excel_path}\n错误详情: {error_msg}"
        ) from error

    try:
        for sheet_name in sheet_names:
            if sheet_name not in wb.sheetnames:
                if warn:
                    warn(f"Sheet '{sheet_name}' 不存在，跳过")
                continue

            ws = wb[sheet_name]
            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            header_norm = [normalize_cell_text(x).replace(" ", "").lower() for x in header]

            func_idx = None
            kw_idx = None
            capl_idx = None
            remark_idx = None

            for i, cell_text in enumerate(header_norm):
                if func_idx is None and ("函数" in cell_text or cell_text == "func"):
                    func_idx = i
                if kw_idx is None and ("关键字" in cell_text or cell_text == "keyword"):
                    kw_idx = i
                if capl_idx is None and ("capl函数" in cell_text or "capl" in cell_text):
                    capl_idx = i
                if remark_idx is None and ("备注" in cell_text or cell_text == "remark"):
                    remark_idx = i

            if kw_idx is None:
                kw_idx = 0 if len(header_norm) > 0 else None
            if capl_idx is None:
                capl_idx = 1 if len(header_norm) > 1 else None
            if capl_idx is None:
                if warn:
                    warn(f"Sheet '{sheet_name}' 缺少 CAPL函数 列，跳过")
                continue

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                func_name = (
                    normalize_cell_text(row[func_idx])
                    if func_idx is not None and func_idx < len(row)
                    else ""
                )
                keyword = (
                    normalize_cell_text(row[kw_idx])
                    if kw_idx is not None and kw_idx < len(row)
                    else ""
                )
                capl_func = (
                    normalize_cell_text(row[capl_idx])
                    if capl_idx is not None and capl_idx < len(row)
                    else ""
                )
                remark = (
                    normalize_cell_text(row[remark_idx])
                    if remark_idx is not None and remark_idx < len(row)
                    else ""
                )

                if not capl_func:
                    continue

                spec = KeywordSpec(
                    func_name=func_name,
                    keyword=keyword,
                    capl_func=capl_func,
                    remark=remark,
                    source_sheet=sheet_name,
                )
                full_key = spec.full_key
                if not full_key:
                    if warn:
                        warn(f"Sheet '{sheet_name}' 第{row_idx}行：函数和关键字都为空，跳过")
                    continue

                lowered_key = full_key.lower()
                if lowered_key in specs and warn:
                    warn(f"重复的键: {full_key} (sheet={sheet_name}, 行{row_idx})")
                specs[lowered_key] = spec
    finally:
        wb.close()

    return specs
