#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
configuration.xlsx 枚举翻译（供 CAN/CIN 生成器复用）

功能概览：
  1. 从 Configuration.txt 的 [CONFIG_ENUM] 段读取 Inputs，解析 configuration.xlsx 路径及 Sheet 列表。
  2. 打开 configuration.xlsx，按 Name / Values 列构建「Name -> 枚举文本->数值」映射（不做 Name->Path 替换）。
  3. 专用于关键字 Set_Config：对步骤参数 [name, value...] 做枚举翻译（value 文本 -> Values 中左侧数值），name 不替换。
  4. 纯数值、含表达式符号 '><=()' 的 value 不翻译，直接透传；与 IOMAPPING 一致：只要配置了 Inputs 就启用（不检查 Enabled）。
"""

from __future__ import annotations

import os
import re
from dataclasses import dataclass
from typing import Dict, List, Optional

from openpyxl import load_workbook

from utils.excel_io import split_input_lines


class ConfigEnumParseError(Exception):
    """configuration.xlsx 枚举翻译失败时抛出，供上层生成注释行或报错。"""


_RE_NUMERIC = re.compile(r"^\s*[-+]?\d+(?:\.\d+)?\s*$")
_RE_ENGLISH_PHRASE = re.compile(r"^[A-Za-z]+(?:\s+[A-Za-z]+)*$")
_EXPR_CHARS = set("><=()")
_COLON_CHARS = (":", "\uFF1A")


def _find_colon(s: str, start: int) -> int:
    """在 s[start:] 中查找第一个冒号（半角/全角）位置。参数: s — 字符串；start — 起始下标。返回: 索引，无则 -1。"""
    candidates = [s.find(c, start) for c in _COLON_CHARS]
    candidates = [p for p in candidates if p >= 0]
    return min(candidates) if candidates else -1


def _norm_key(s: str) -> str:
    """Name/键规范化：去首尾空白、转小写。参数: s — 原始字符串。返回: str。"""
    return str(s).strip().casefold()


def _is_numeric(s: str) -> bool:
    """判断是否为十进制数值。参数: s — 待判断字符串。返回: bool。"""
    return bool(s is not None and _RE_NUMERIC.match(str(s)))


def _has_expr_chars(s: str) -> bool:
    """判断是否含表达式符号 ><=()，此类值不翻译。参数: s — 待判断字符串。返回: bool。"""
    if s is None:
        return False
    return any((ch in _EXPR_CHARS) for ch in str(s))


def _parse_values_cell(values_cell: str) -> Dict[str, str]:
    """解析 Values 单元格为「翻译前(右)->翻译后(左)」映射。参数: values_cell — 单元格字符串。返回: 规范 key -> 左侧值。"""
    if values_cell is None:
        return {}
    text = str(values_cell).strip()
    if not text:
        return {}

    mapping: Dict[str, str] = {}
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue

        i = 0
        n = len(line)
        while i < n:
            j = _find_colon(line, i)
            if j < 0:
                break
            left = line[i:j].strip()
            if not left:
                i = j + 1
                continue
            k = j + 1
            while k < n and line[k].isspace():
                k += 1
            start_right = k
            next_pair_pos = None
            m = re.search(r"\s+\S+\s*[:\uFF1A]", line[start_right:])
            if m:
                next_pair_pos = start_right + m.start()
            if next_pair_pos is None:
                right = line[start_right:].strip()
                i = n
            else:
                right = line[start_right:next_pair_pos].strip()
                i = next_pair_pos
                while i < n and line[i].isspace():
                    i += 1
            if not right:
                continue
            mapping[_norm_key(right)] = left.strip()

    return mapping


@dataclass
class ConfigEnumContext:
    """配置枚举上下文：仅持有 name_to_values，供 Set_Config 步骤参数翻译用。"""

    name_to_values: Dict[str, Dict[str, str]]

    def translate_args(self, args: List[str]) -> List[str]:
        """对 Set_Config 步骤参数做枚举翻译：args[0] 为 Name 不替换，后续 value 按 Values 映射。参数: args — [name, value...]。返回: 翻译后的参数列表。"""
        if not args:
            raise ConfigEnumParseError("参数为空")
        raw_name = str(args[0]).strip()
        if not raw_name:
            raise ConfigEnumParseError("Name 为空")

        name_key = _norm_key(raw_name)
        values_map = self.name_to_values.get(name_key)
        values_empty = False
        if not values_map:
            if name_key not in self.name_to_values:
                raise ConfigEnumParseError(f"Name 未找到: {raw_name}")
            values_empty = True
            values_map = {}

        out: List[str] = [raw_name]
        if len(args) == 1:
            return out

        rest_tokens: List[str] = []
        for raw in args[1:]:
            t = str(raw).strip()
            if t:
                rest_tokens.append(t)
        rest_str = " ".join(rest_tokens).strip()
        if rest_str:
            if _is_numeric(rest_str):
                out.append(rest_str)
                return out
            if _has_expr_chars(rest_str):
                out.append(rest_str)
                return out
            if values_empty:
                raise ConfigEnumParseError(f"Values 为空: Name={raw_name}, Value={rest_str}")
            if _RE_ENGLISH_PHRASE.match(rest_str):
                k_all = _norm_key(rest_str)
                if k_all in values_map:
                    out.append(values_map[k_all])
                    return out
                raise ConfigEnumParseError(f"Values 未匹配: Name={raw_name}, Value={rest_str}")
            k_all = _norm_key(rest_str)
            if k_all in values_map:
                out.append(values_map[k_all])
                return out

        i = 1
        while i < len(args):
            tok = str(args[i]).strip()
            if not tok:
                i += 1
                continue

            if _is_numeric(tok):
                out.append(tok)
                i += 1
                continue

            if i + 1 < len(args):
                tok2 = str(args[i + 1]).strip()
                two = f"{tok} {tok2}".strip()
                two_key = _norm_key(two)
                if two_key in values_map:
                    out.append(values_map[two_key])
                    i += 2
                    continue

            one_key = _norm_key(tok)
            if one_key in values_map:
                out.append(values_map[one_key])
                i += 1
                continue

            raise ConfigEnumParseError(f"Values 未匹配: Name={raw_name}, Value={tok}")

        return out


def _get_config_enum_inputs_text(config, domain: str) -> str:
    """从配置中按域读取 CONFIG_ENUM 的 Inputs 文本。参数: config — 配置对象；domain — 域（如 LR_REAR，兼容全局 CONFIG_ENUM）。返回: Inputs 字符串。"""
    section_candidates = [f"{domain}_CONFIG_ENUM"] if domain and domain != "LR_REAR" else ["CONFIG_ENUM"]
    for section in section_candidates:
        if not config.has_section(section):
            continue
        inputs_text = (
            config.get(section, "Inputs", fallback="")
            or config.get(section, "inputs", fallback="")
        ).strip("\n")
        if inputs_text.strip():
            return inputs_text
    return ""


def load_config_enum_from_config(
    config,
    base_dir: Optional[str] = None,
    config_path: Optional[str] = None,
    domain: str = "LR_REAR",
) -> Optional[ConfigEnumContext]:
    """从 Configuration.txt 的 [CONFIG_ENUM] 段加载 configuration.xlsx，构建 Name->Values 枚举上下文。
    参数: config — 配置对象；base_dir — 工程根目录；config_path — 配置文件路径；domain — 域。
    返回: ConfigEnumContext 或 None（未配置 Inputs 时）。
    """
    inputs_text = _get_config_enum_inputs_text(config, domain)
    inputs = split_input_lines(inputs_text)
    if not inputs:
        return None

    if config_path:
        config_dir = os.path.dirname(os.path.abspath(config_path))
    elif base_dir:
        config_dir = base_dir
    else:
        config_dir = os.getcwd()

    name_to_values: Dict[str, Dict[str, str]] = {}

    for rel_path, sheets_str in inputs:
        excel_path = rel_path.strip()
        excel_path_for_check = excel_path.replace("/", os.sep)
        if not os.path.isabs(excel_path_for_check):
            excel_path_for_check = os.path.abspath(os.path.join(config_dir, excel_path_for_check))
        excel_path_for_check = os.path.normpath(excel_path_for_check)

        if not os.path.exists(excel_path_for_check):
            try:
                excel_path_utf8 = excel_path_for_check.encode("utf-8").decode("utf-8")
                if os.path.exists(excel_path_utf8):
                    excel_path_for_check = excel_path_utf8
                else:
                    raise ConfigEnumParseError(f"找不到 configuration.xlsx: {excel_path_for_check}")
            except Exception:
                raise ConfigEnumParseError(f"找不到 configuration.xlsx: {excel_path_for_check}")

        excel_path = excel_path_for_check.replace("\\", "/")

        try:
            wb = load_workbook(excel_path, data_only=True, read_only=True)
        except Exception as e:
            error_msg = str(e)
            if (
                "decompressing" in error_msg.lower()
                or "incorrect header" in error_msg.lower()
                or "badzipfile" in error_msg.lower()
            ):
                raise ConfigEnumParseError(
                    f"Configuration Excel 文件格式错误或文件已损坏: {excel_path}\n"
                    f"错误详情: {error_msg}\n"
                    f"请检查文件是否是有效的 Excel 文件（.xlsx 格式）"
                )
            raise ConfigEnumParseError(
                f"无法读取 Configuration Excel 文件: {excel_path}\n错误详情: {error_msg}"
            )

        if not sheets_str or sheets_str.strip() == "*" or sheets_str.strip() == "":
            sheets = list(wb.sheetnames)
        else:
            sheets = [s.strip() for s in sheets_str.split(",") if s.strip()]
            sheets = [s for s in sheets if s in wb.sheetnames]
            if not sheets:
                sheets = list(wb.sheetnames)

        for sheet in sheets:
            ws = wb[sheet]
            try:
                header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            except StopIteration:
                continue
            headers = [str(h).strip() if h is not None else "" for h in header]

            name_idx = None
            values_idx = None
            for i, h in enumerate(headers):
                if h == "Name":
                    name_idx = i
                if h == "Values":
                    values_idx = i
            if name_idx is None or values_idx is None:
                continue

            for row in ws.iter_rows(min_row=2, values_only=True):
                name_cell = row[name_idx] if len(row) > name_idx else None
                values_cell = row[values_idx] if len(row) > values_idx else None

                name_s = str(name_cell).strip() if name_cell is not None else ""
                values_s = str(values_cell).strip() if values_cell is not None else ""

                if not name_s:
                    continue
                enum_map = _parse_values_cell(values_s) if values_s else {}
                name_to_values[_norm_key(name_s)] = enum_map

    if not name_to_values:
        return None

    return ConfigEnumContext(name_to_values=name_to_values)


__all__ = [
    "load_config_enum_from_config",
    "ConfigEnumContext",
    "ConfigEnumParseError",
]
