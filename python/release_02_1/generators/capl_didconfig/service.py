#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""DIDConfig 生成调度 service。不再依赖 DIDConfigLegacyHooks，通过 .runtime 委托根脚本与公共模块。"""

from __future__ import annotations

import os
import sys

from openpyxl import load_workbook

from utils.path_utils import resolve_target_subdir

from . import runtime as _rt


class DIDConfigGeneratorService:
    """接管 DIDConfig 旧版主编排流程的 service。"""

    def run_legacy_pipeline(self):
        base_dir = _rt.resolve_base_dir()
        gconfig = _rt.load_runtime(base_dir)
        if gconfig is None:
            return

        log_mgr, logger, old_stdout, old_stderr = _rt.init_logging(base_dir)
        progress_level = _rt.get_progress_level()
        try:
            cfg = gconfig.raw_config
            if not cfg.has_section("DID_CONFIG"):
                msg = "未配置 DID_Config 配置节 [DID_CONFIG]"
                print(f"错误: {msg}")
                logger.error(msg)
                # 抛异常，交由 TaskService 决定是“跳过”还是失败，避免前端误认为已生成
                raise ValueError(msg)

            excel_rel_path = cfg.get("DID_CONFIG", "input_excel", fallback=None) or cfg.get(
                "DID_CONFIG", "Input_Excel", fallback=None
            )
            if not excel_rel_path:
                msg = "未配置 DID_Config 配置表：配置文件中未找到 DID_CONFIG.input_excel 或 DID_CONFIG.Input_Excel"
                print(f"错误: {msg}")
                logger.error(msg)
                # 抛异常以便上层按“未配置时跳过”处理，而不是静默 return
                raise ValueError(msg)

            output_name = (
                gconfig.get_fixed("didconfig_output_filename")
                or cfg.get("DID_CONFIG", "output_filename", fallback=None)
                or cfg.get("DID_CONFIG", "Output_FileName", fallback="DIDConfig.txt")
            )
            output_dir_rel = cfg.get("DID_CONFIG", "output_dir", fallback=None) or cfg.get(
                "DID_CONFIG", "Output_Dir", fallback=None
            )
            if not output_dir_rel:
                msg = "配置文件中未找到 DID_CONFIG.output_dir 或 DID_CONFIG.Output_Dir"
                print(f"错误: {msg}")
                logger.error(msg)
                raise ValueError(msg)

            config_dir = gconfig.config_dir
            excel_rel_path_normalized = excel_rel_path.replace("/", os.sep)
            if os.path.isabs(excel_rel_path_normalized):
                excel_path = excel_rel_path_normalized
            else:
                excel_path = os.path.abspath(os.path.join(config_dir, excel_rel_path_normalized))
            excel_path = os.path.normpath(excel_path)

            output_dir_rel_normalized = output_dir_rel.replace("/", os.sep)
            if os.path.isabs(output_dir_rel_normalized):
                raw_output_dir = output_dir_rel_normalized
            else:
                raw_output_dir = os.path.abspath(os.path.join(config_dir, output_dir_rel_normalized))
            raw_output_dir = os.path.normpath(raw_output_dir)

            # 自动寻找或创建 Configuration 文件夹（与 DIDInfo、UART 等模块一致）
            try:
                output_dir = resolve_target_subdir(base_dir, raw_output_dir, "Configuration")
            except Exception as e:
                logger.error(f"无法定位输出目录: {e}")
                return
            output_path = os.path.join(output_dir, output_name)

            if not os.path.exists(excel_path):
                print(f"错误: 找不到 Excel 文件 {excel_path}")
                return

            try:
                wb = load_workbook(excel_path, data_only=True)
            except Exception as e:
                error_msg = str(e)
                if (
                    "decompressing" in error_msg.lower()
                    or "incorrect header" in error_msg.lower()
                    or "badzipfile" in error_msg.lower()
                ):
                    raise ValueError(
                        f"DIDConfig Excel 文件格式错误或文件已损坏: {excel_path}\n"
                        f"错误详情: {error_msg}\n"
                        "请检查文件是否是有效的 Excel 文件（.xlsx 格式）"
                    )
                raise ValueError(f"无法读取 DIDConfig Excel 文件: {excel_path}\n错误详情: {error_msg}")

            excel_name = os.path.basename(excel_path)
            output_content: list[str] = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                logger.log(progress_level, f"解析 Excel 文件: {excel_name} sheet={sheet_name}")
                header_row, cols, missing = _rt.find_header_row_and_col_indices(
                    ws,
                    {
                        "name": ["name", "field_subdataname", "subdataname"],
                        "byte": ["byte", "field_byteposition", "byteposition"],
                        "bit": ["bit", "field_bitposition", "bitposition"],
                    },
                    max_scan_rows=30,
                )
                if header_row < 0 or missing:
                    print(f"错误: Excel={excel_name} sheet='{sheet_name}' 缺少必需列: {', '.join(missing)}，跳过该 sheet。")
                    continue

                parsed_rows: list[tuple[int, str, int, str]] = []
                for r in range(header_row + 1, ws.max_row + 1):
                    name_v = _rt.merged_cell_value(ws, r, cols["name"])
                    byte_v = _rt.merged_cell_value(ws, r, cols["byte"])
                    bit_v = _rt.merged_cell_value(ws, r, cols["bit"])

                    name_s = _rt.norm_str(name_v)
                    byte_s = _rt.norm_str(byte_v)
                    bit_s = _rt.norm_str(bit_v)

                    if not name_s and not byte_s and not bit_s:
                        continue

                    shown_name = name_s if name_s else "<空>"
                    if not byte_s and not bit_s:
                        print(f"错误: Excel={excel_name} sheet='{sheet_name}' 行 {r} Name='{shown_name}' Byte/Bit 均为空，跳过该行。")
                        continue
                    if not byte_s:
                        print(f"错误: Excel={excel_name} sheet='{sheet_name}' 行 {r} Name='{shown_name}' Byte 为空，跳过该行。")
                        continue
                    if not bit_s:
                        print(f"错误: Excel={excel_name} sheet='{sheet_name}' 行 {r} Name='{shown_name}' Bit 为空，跳过该行。")
                        continue
                    if not name_s:
                        print(f"错误: Excel={excel_name} sheet='{sheet_name}' 行 {r} Name 为空（Byte={byte_s}, Bit={bit_s}），跳过该行。")
                        continue

                    try:
                        byte_int = int(str(byte_v).strip())
                    except Exception:
                        print(f"错误: Excel={excel_name} sheet='{sheet_name}' 行 {r} Name='{shown_name}' Byte 无法解析为整数: '{byte_s}'，跳过该行。")
                        continue

                    parsed_rows.append((r, name_s, byte_int, bit_s))

                if not parsed_rows:
                    continue

                did_id = sheet_name if sheet_name.startswith("0x") else f"0x{sheet_name}"
                output_content.append(f"[{did_id}]")
                max_byte = max((b for (_, _, b, _) in parsed_rows), default=None)
                did_length = (max_byte + 1) if isinstance(max_byte, int) else 0
                output_content.append(f"DIDLength:{did_length};//DID数据长度BYTE")

                for excel_row, name, byte, bit_raw in parsed_rows:
                    bit_s = _rt.norm_str(bit_raw).lower()
                    if bit_s == "all":
                        bit_pos = 0
                        field_len = 8
                    else:
                        try:
                            parts = bit_s.split("-")
                            if len(parts) == 1:
                                bit_pos = int(parts[0])
                                field_len = 1
                            elif len(parts) == 2:
                                start = int(parts[0])
                                end = int(parts[1])
                                bit_pos = start
                                field_len = max(1, end - start + 1)
                            else:
                                raise ValueError(f"Bit 格式不支持: '{bit_raw}'")
                        except Exception:
                            print(f"错误: Excel={excel_name} sheet='{sheet_name}' 行 {excel_row} Name='{name}' Bit 无法解析: '{bit_raw}'，跳过该行。")
                            continue

                    output_content.append(f"Field_subDataName:{name};//字段名称")
                    output_content.append(f"Field_BytePosition:{byte};//字段起始byte")
                    output_content.append(f"Field_bitPosition:{bit_pos};//字段起始bit;")
                    output_content.append(f"Field_FieldLength:{field_len};//字段长度，单位bit;")
                    output_content.append("Field_SortOrder:0;//排序方式0是motorola,1是intel")
                    output_content.append("Field_Data:0x01;//字段数据")
                    output_content.append("")

            with open(output_path, "w", encoding="utf-8") as f:
                f.write("\n".join(output_content))

            print(f"成功生成: {output_path}")
            logger.log(progress_level, f"成功生成: {output_path}")
        finally:
            sys.stdout = old_stdout
            sys.stderr = old_stderr
            log_mgr.clear()
