#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""CIN 入口运行期辅助。"""

from __future__ import annotations

import os

from infra.filesystem.pathing import RuntimePathResolver, resolve_target_subdir
from openpyxl import load_workbook

from core.generator_config import GeneratorConfig
from core.mapping_context import MappingContext
from services.config_constants import (
    DEFAULT_DOMAIN_LR_REAR,
    OPTION_CIN_INPUT_EXCEL_CANDIDATES,
    OPTION_OUTPUT_DIR_CANDIDATES,
    SECTION_LR_REAR,
    SECTION_PATHS,
)
from .constants import (
    DEFAULT_CIN_OUTPUT_DIR,
    DEFAULT_CIN_OUTPUT_FILENAME,
    DEFAULT_KEYWORD_SHEET_NAME,
)
class CINEntrypointSupport:
    """收拢 CIN 入口阶段的路径、配置与上下文初始化。"""

    @staticmethod
    def resolve_base_dir() -> str:
        """返回项目根目录（含主配置 / 固定配置标记文件的目录）。"""
        return RuntimePathResolver.resolve_base_dir(__file__)

    @staticmethod
    def load_runtime_config(base_dir: str) -> dict:
        gconfig = GeneratorConfig(base_dir).load()
        config_path = RuntimePathResolver.resolve_config_path(base_dir, gconfig.config_path)
        if not os.path.exists(config_path):
            raise FileNotFoundError(f"未找到配置文件: {config_path}")

        cfg = gconfig.raw_config
        input_excel_file = gconfig.get_first(
            [
                (SECTION_LR_REAR, option_name) for option_name in OPTION_CIN_INPUT_EXCEL_CANDIDATES
            ] + [
                (SECTION_PATHS, option_name) for option_name in reversed(OPTION_CIN_INPUT_EXCEL_CANDIDATES)
            ]
        )
        input_sheet = gconfig.get_first(
            [
                (SECTION_LR_REAR, "Cin_Input_Sheet"),
                (SECTION_LR_REAR, "cin_input_sheet"),
                (SECTION_PATHS, "Cin_Input_Sheet"),
                (SECTION_PATHS, "cin_input_sheet"),
            ],
            fallback="",
        )
        mapping_excel_file = gconfig.get_fixed("cin_mapping_excel") or gconfig.get_fixed(
            "unified_mapping_excel"
        )
        if not mapping_excel_file:
            mapping_excel_file = gconfig.get(
                SECTION_PATHS,
                "Cin_Mapping_Excel",
                fallback="",
            )
        if not mapping_excel_file or not str(mapping_excel_file).strip():
            raise ValueError("未配置 cin_mapping_excel / unified_mapping_excel / Cin_Mapping_Excel")

        cin_mapping_sheet = gconfig.get_fixed("cin_mapping_sheet") or gconfig.get(
            SECTION_PATHS, "Cin_Mapping_Sheet", fallback=""
        )
        sheet_names_str = gconfig.get_fixed("mapping_sheets") or gconfig.get(
            SECTION_PATHS,
            "Mapping_Sheets",
            fallback="",
        )
        if (
            cin_mapping_sheet
            and str(cin_mapping_sheet).strip()
            and str(cin_mapping_sheet).strip() != DEFAULT_KEYWORD_SHEET_NAME
        ):
            sheet_names_str = str(cin_mapping_sheet).strip()

        output_dir_cin = gconfig.get_first(
            [
                (SECTION_LR_REAR, "Output_Dir_Cin"),
                (SECTION_LR_REAR, "output_dir_cin"),
                (SECTION_PATHS, "Output_Dir_Cin"),
                (SECTION_PATHS, "output_dir_cin"),
            ]
        )
        output_dir = gconfig.get_first(
            [
                (SECTION_LR_REAR, option_name) for option_name in OPTION_OUTPUT_DIR_CANDIDATES
            ] + [
                (SECTION_PATHS, option_name) for option_name in OPTION_OUTPUT_DIR_CANDIDATES
            ]
        )
        if output_dir_cin and str(output_dir_cin).strip():
            output_dir = output_dir_cin
        if not output_dir or not str(output_dir).strip():
            output_dir = DEFAULT_CIN_OUTPUT_DIR

        output_cin_filename = (
            gconfig.get_fixed("cin_output_filename") or DEFAULT_CIN_OUTPUT_FILENAME
        )

        if not input_excel_file:
            raise ValueError("未配置 Cin_Input_Excel 或 cin_input_excel")
        input_excel_path = (
            input_excel_file
            if os.path.isabs(input_excel_file)
            else os.path.join(base_dir, input_excel_file)
        )
        if mapping_excel_file.startswith("./"):
            mapping_excel_file = mapping_excel_file[2:]
        mapping_excel_path = (
            mapping_excel_file
            if os.path.isabs(mapping_excel_file)
            else os.path.join(base_dir, mapping_excel_file)
        )
        output_dir = resolve_target_subdir(base_dir, output_dir, "TESTmode")

        return {
            "config_path": config_path,
            "cfg": cfg,
            "input_sheet": input_sheet,
            "input_excel_path": input_excel_path,
            "mapping_excel_path": mapping_excel_path,
            "sheet_names_str": sheet_names_str,
            "output_dir": output_dir,
            "output_cin_filename": output_cin_filename,
        }

    @staticmethod
    def detect_sheet_title(input_excel_path: str, input_sheet: str | None) -> str:
        try:
            wb_temp = load_workbook(input_excel_path, data_only=True, read_only=True)
            ws_temp = wb_temp.active
            if input_sheet and str(input_sheet).strip():
                sn = str(input_sheet).strip()
                if sn in wb_temp.sheetnames:
                    ws_temp = wb_temp[sn]
            sheet_title = ws_temp.title
            wb_temp.close()
            return sheet_title
        except Exception:
            return input_sheet if input_sheet else "未知"

    @staticmethod
    def load_mapping_context(cfg, base_dir: str, config_path: str, domain: str = DEFAULT_DOMAIN_LR_REAR):
        mapping_ctx = MappingContext.from_config(
            cfg,
            base_dir=base_dir,
            config_path=config_path,
            domain=domain,
        )
        return mapping_ctx.io_mapping, mapping_ctx.config_enum
