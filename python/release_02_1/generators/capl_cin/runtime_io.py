#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CIN 生成运行期 IO 与步骤解析。

从根目录脚本迁入：load_keyword_specs、read_clib_steps、render_step_lines、generate_content；
以及 reset_runtime_state、setup_generator_logger、load_mapping_context，供根目录编排与兼容入口使用。
"""

from __future__ import annotations

import os
import sys
from typing import Any, Callable, Optional, Tuple

from openpyxl import load_workbook

from core.generator_logging import GeneratorLogger
from core.mapping_context import MappingContext

from .constants import CASEID_LOG_PATTERNS
from .runtime import CINEntrypointSupport

from core.error_module import ErrorModuleResolver
from core.parser import KeywordMatchError, StepSyntaxError, parse_step_line
from core.step_error_detail import StepErrorDetailBuilder, format_step_error_lines
from core.translator import (
    ConfigEnumParseError,
    IOMappingParseError,
    load_keyword_specs_from_excel,
)
from core.common.name_sanitize import sanitize_clib_name


def load_keyword_specs(
    excel_path: str,
    sheet_names: list[str],
    *,
    warn: Optional[Callable[[str], None]] = None,
) -> dict:
    """读取关键字-CAPL 映射表。"""
    return load_keyword_specs_from_excel(
        excel_path,
        sheet_names,
        warn=warn or (lambda msg: None),
    )


def read_clib_steps(excel_path: str, clib_sheet: Optional[str] = None) -> tuple[str, list]:
    """
    从 Clib Excel 读取 Name/Step 列，按 Name 聚合步骤。
    返回 (sheet_title, [(name_str, [(step_text, excel_row_num), ...]), ...])。
    """
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"找不到 Clib Excel 文件: {excel_path}")

    try:
        wb = load_workbook(excel_path, data_only=True)
    except Exception as e:
        err = str(e).lower()
        if "decompressing" in err or "incorrect header" in err or "badzipfile" in err:
            raise ValueError(
                f"Excel 文件格式错误或文件已损坏: {excel_path}\n错误详情: {e}\n"
                "请检查文件是否是有效的 Excel 文件（.xlsx 格式）"
            ) from e
        raise ValueError(f"无法读取 Excel 文件: {excel_path}\n错误详情: {e}") from e

    ws = wb.active
    if clib_sheet and str(clib_sheet).strip():
        sn = str(clib_sheet).strip()
        if sn in wb.sheetnames:
            ws = wb[sn]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    header = [str(h).strip() if h is not None else "" for h in header_row]
    if not header:
        raise ValueError(f"[cin] 错误: 工作表 '{ws.title}' 没有表头")

    name_idx = step_idx = None
    for i, h in enumerate(header):
        h_lower = h.lower()
        if "name" in h_lower and name_idx is None and "project" not in h_lower:
            name_idx = i
        if "step" in h_lower and step_idx is None:
            step_idx = i
    if name_idx is None:
        name_idx = 2 if len(header) >= 2 and "project" in header[1].lower() else 1
    if step_idx is None:
        step_idx = (name_idx + 1) if name_idx is not None else 2

    ordered: list = []
    seen: dict = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        name = row[name_idx] if len(row) > name_idx else None
        step_block = row[step_idx] if len(row) > step_idx else None
        if not name or not step_block:
            continue
        name_str = str(name).strip()
        if not name_str:
            continue
        if name_str not in seen:
            steps = []
            seen[name_str] = steps
            ordered.append((name_str, steps))
        else:
            steps = seen[name_str]
        for raw_line in str(step_block).splitlines():
            s = str(raw_line).rstrip("\n")
            if s.strip():
                steps.append((s, row_idx))

    return ws.title, ordered


def _is_numeric(s: str) -> bool:
    try:
        float(s)
        return True
    except ValueError:
        return False


def _apply_default_param_parsing(args: list[str]) -> list[str]:
    parsed = []
    for arg in args:
        parsed.append(arg if _is_numeric(arg) else f'"{arg}"')
    return parsed


# 模块内保存上一次解析错误，供 render_step_lines 中 build_detail 使用
_last_parse_error: dict = {"type": None, "reason": ""}


def _parse_step_line_cin(
    line: str,
    keyword_specs: dict,
    *,
    io_mapping_ctx: Any,
    config_enum_ctx: Any,
    logger: Optional[Any],
    name: Optional[str] = None,
) -> Optional[tuple[list[str], str]]:
    """CIN 模式下一行步骤解析，返回 (code_lines, original_line_full) 或 None。"""
    global _last_parse_error
    _last_parse_error["type"] = None
    _last_parse_error["reason"] = ""
    original = line.strip()
    if not original or original.startswith("//"):
        return None

    try:
        result = parse_step_line(
            original,
            keyword_specs,
            mode="cin",
            io_mapping_ctx=io_mapping_ctx,
            config_enum_ctx=config_enum_ctx,
            sanitize_clib_name=sanitize_clib_name,
            default_param_parser=_apply_default_param_parsing,
        )
        if result is None:
            return None
        return (result.code_lines, result.original_line_full)
    except IOMappingParseError as e:
        _last_parse_error["type"] = "iomapping_conflict" if "CONFLICT" in str(e) else "iomapping"
        _last_parse_error["reason"] = str(e)
        if logger:
            step_text = original.strip()
            reason = str(e)
            fail_text = reason if reason.startswith("IO_mapping") else f"IO_mapping 表中{reason}"
            name_part = f"Clib_Name：{name}" if name else "Clib_Name：未知"
            err_mod = ErrorModuleResolver.resolve(fail_text)
            logger.error(f"错误模块【{err_mod}】 {name_part} 用例步骤：{step_text}  原因：{fail_text}")
        return None
    except ConfigEnumParseError as e:
        _last_parse_error["type"] = "config_enum"
        _last_parse_error["reason"] = str(e)
        if logger:
            step_text = original.strip()
            reason = str(e)
            fail_text = reason if reason.startswith("CONFIG_ENUM") else f"CONFIG_ENUM 表中{reason}"
            name_part = f"Clib_Name：{name}" if name else "Clib_Name：未知"
            err_mod = ErrorModuleResolver.resolve(fail_text)
            logger.error(f"错误模块【{err_mod}】 {name_part} 用例步骤：{step_text}  原因：{fail_text}")
        return None
    except (KeywordMatchError, StepSyntaxError) as e:
        _last_parse_error["type"] = "keyword" if isinstance(e, KeywordMatchError) else "syntax"
        _last_parse_error["reason"] = str(e)
        if logger and isinstance(e, StepSyntaxError):
            fail_text = f"步骤语法错误: {e}"
            name_part = f"Clib_Name：{name}" if name else "Clib_Name：未知"
            err_mod = ErrorModuleResolver.resolve(fail_text)
            logger.error(f"错误模块【{err_mod}】 {name_part} 用例步骤：{original.strip()}  原因：{fail_text}")
        return None
    return None


def render_step_lines(
    raw_line: str,
    keyword_specs: dict,
    *,
    io_mapping_ctx: Any,
    config_enum_ctx: Any,
    logger: Optional[Any] = None,
    source_id: Optional[str] = None,
    excel_name: Optional[str] = None,
    sheet_name: Optional[str] = None,
    name: Optional[str] = None,
    excel_row_num: Optional[int] = None,
) -> Optional[list[str]]:
    """
    将一行 Step 转为 0..N 行 CIN CAPL 代码（含 //测试步骤 注释与 [idx/total]）；
    解析失败时生成错误占位行。
    """
    original_line = str(raw_line).strip()
    if original_line.startswith("//"):
        return None

    result = _parse_step_line_cin(
        original_line,
        keyword_specs,
        io_mapping_ctx=io_mapping_ctx,
        config_enum_ctx=config_enum_ctx,
        logger=logger,
        name=name,
    )

    if result is None:
        err_type = _last_parse_error.get("type")
        err_reason = _last_parse_error.get("reason", "")
        if err_type == "iomapping_conflict":
            return None
        error_detail = StepErrorDetailBuilder.build_detail(
            err_type or "unknown",
            err_reason,
            original_line,
            keyword_specs,
        )
        return format_step_error_lines(original_line, error_detail, role_prefix="测试步骤")

    code_lines, original_line_full = result
    if not code_lines:
        return None

    out = []
    for idx, code in enumerate(code_lines, start=1):
        suffix = f" [{idx}/{len(code_lines)}]" if len(code_lines) > 1 else ""
        out.append(code.rstrip() + f" //测试步骤 {original_line_full}" + suffix)
    return out


def generate_content(
    ordered_func_steps: list,
    include_files: Optional[list[str]] = None,
) -> tuple[str, list]:
    """
    将 (export_func_name, steps) 列表拼成 .cin 全文。
    返回 (cin_content_str, error_records)，error_records 为 [(func_name, teststep_content, teststepfail_content), ...]。
    """
    lines = []
    error_records: list = []
    lines.append("/*@!Encoding:936*/")
    lines.append("includes")
    lines.append("{")
    if include_files:
        for inc in include_files:
            lines.append(f'  #include "{inc}"')
    else:
        lines.append("  ")
    lines.append("}")
    lines.append("")
    lines.append("variables")
    lines.append("{")
    lines.append("  ")
    lines.append("}")
    lines.append("")

    for export_func_name, steps in ordered_func_steps:
        lines.append(f"export void {export_func_name}()")
        lines.append("{")
        teststep_content = None
        for step in steps or []:
            lines.append(step)
            if step.strip().startswith("teststep("):
                teststep_content = step.strip()
            elif step.strip().startswith("teststepfail(") and teststep_content is not None:
                error_records.append((export_func_name, teststep_content, step.strip()))
                teststep_content = None
        lines.append("}")
        lines.append("")

    return "\n".join(lines), error_records


# ==================== 运行时上下文与编排用 API（去脚本化） ====================

_io_mapping_ctx: Any = None
_config_enum_ctx: Any = None


def reset_runtime_state() -> None:
    """初始化/重置全局上下文状态，避免跨任务串状态。"""
    global _io_mapping_ctx, _config_enum_ctx
    _io_mapping_ctx = None
    _config_enum_ctx = None


def setup_generator_logger(base_dir: str) -> GeneratorLogger:
    """步骤 ①：初始化 CIN 主日志管理器（generate_cin_from_excel.log），含 CASEID 过滤。"""
    try:
        from infra.logger import ProgressFormatter, SubstringFilter
    except ImportError:
        from utils.logger import ProgressFormatter, SubstringFilter
    return GeneratorLogger(
        base_dir,
        log_basename="generate_cin_from_excel.log",
        logger_name="generate_cin_from_excel",
        formatter_factory=lambda s: ProgressFormatter(s),
        file_filters=[SubstringFilter(CASEID_LOG_PATTERNS, include=False)],
    )


def load_mapping_context(
    cfg: Any, base_dir: str, config_path: str, domain: str = "LR_REAR"
) -> Tuple[Any, Any]:
    """步骤 ②：按 domain 加载 io_mapping 与 Configuration 枚举上下文；写入模块级供 legacy 使用。"""
    global _io_mapping_ctx, _config_enum_ctx
    _io_mapping_ctx, _config_enum_ctx = CINEntrypointSupport.load_mapping_context(
        cfg, base_dir=base_dir, config_path=config_path, domain=domain
    )
    return _io_mapping_ctx, _config_enum_ctx


def legacy_read_clib_steps(excel_path: str, clib_sheet: Optional[str] = None) -> tuple:
    """兼容入口：从 Clib Excel 读取 Name/Step，与 read_clib_steps 行为一致。"""
    return read_clib_steps(excel_path, clib_sheet=clib_sheet)
