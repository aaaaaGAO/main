#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""CAN 入口运行期辅助。"""

from __future__ import annotations

import os
import sys

from openpyxl import load_workbook

from core.generator_config import GeneratorConfig
from infra.filesystem.pathing import get_project_root
from utils.path_utils import list_excel_files, resolve_target_subdir


class CANEntrypointSupport:
    """收拢 CAN 入口阶段的路径、配置与兼容辅助。"""

    @staticmethod
    def _domain_candidates(domain: str, *options: str) -> list[tuple[str, str]]:
        pairs = [(domain, option) for option in options]
        if domain == "LR_REAR":
            pairs.extend([("PATHS", option) for option in options])
        return pairs

    @staticmethod
    def parse_bool(raw, default: bool = False) -> bool:
        if raw is None:
            return default
        s = str(raw).strip().lower()
        if s in ("1", "true", "yes", "on"):
            return True
        if s in ("0", "false", "no", "off"):
            return False
        return default

    @staticmethod
    def _resolve_uds_qualifier(
        gconfig: GeneratorConfig,
        base_dir: str,
        domain: str,
        output_dir: str,
    ) -> str:
        # 优先使用各域在 Configuration.txt 中配置的 uds_ecu_qualifier，
        # 这样即使多个域共用同一个 output_dir 也不会互相覆盖。
        cfg_val = gconfig.get(domain, "uds_ecu_qualifier", fallback="").strip()
        if cfg_val:
            return cfg_val

        # 兼容旧流程：若未在配置中显式填写，则尝试从 uds.txt（或自定义 uds_output_filename）中读取。
        fixed = gconfig.fixed_config
        uds_filename = (fixed.get("uds_output_filename") or "uds.txt").strip() or "uds.txt"
        root = output_dir if os.path.isabs(output_dir) else os.path.join(base_dir, output_dir)
        root = os.path.abspath(root)
        config_dir = (
            root if os.path.basename(root).lower() == "configuration" else os.path.join(root, "Configuration")
        )
        uds_path = os.path.join(config_dir, uds_filename)
        if os.path.isfile(uds_path):
            try:
                with open(uds_path, "r", encoding="utf-8", errors="replace") as f:
                    for line in f:
                        if line.strip().lower().startswith("ecu_qualifier="):
                            return line.split("=", 1)[1].strip()
            except Exception:
                pass
        return ""

    @staticmethod
    def build_runtime_paths(gconfig: GeneratorConfig, domain: str = "LR_REAR") -> dict:
        base_dir = gconfig.base_dir

        raw_path = gconfig.get_first(
            CANEntrypointSupport._domain_candidates(
                domain,
                "input_excel",
                "input_excel_dir",
                "Input_Excel_Dir",
                "Input_Excel",
            )
        )
        if not raw_path:
            raise ValueError(f"未配置输入路径：请配置 [{domain}] 的 input_excel。")

        full_path = (
            os.path.join(base_dir, raw_path)
            if not os.path.isabs(raw_path)
            else os.path.normpath(raw_path)
        )
        if os.path.isdir(full_path):
            excel_files = list_excel_files(full_path)
            if not excel_files:
                raise FileNotFoundError(f"文件夹内未找到 Excel 文件: {full_path}")
        else:
            excel_files = [full_path]

        mapping_excel_file = gconfig.get_fixed("mapping_excel") or gconfig.get_fixed(
            "unified_mapping_excel"
        )
        if not mapping_excel_file:
            mapping_excel_file = gconfig.get("PATHS", "Mapping_Excel")
        if not mapping_excel_file:
            raise ValueError(
                "未配置映射表路径，请在 FixedConfig.txt 的 [PATHS] 中配置 mapping_excel 或 unified_mapping_excel。"
            )
        if mapping_excel_file.startswith("./"):
            mapping_excel_file = mapping_excel_file[2:]
        mapping_excel_path = (
            mapping_excel_file
            if os.path.isabs(mapping_excel_file)
            else os.path.join(base_dir, mapping_excel_file)
        )

        output_dir = gconfig.get_first(
            CANEntrypointSupport._domain_candidates(
                domain,
                "Output_Dir_Can",
                "output_dir_can",
                "Output_Dir",
                "output_dir",
            ),
            fallback="./output",
        )
        output_dir = (output_dir or "./output").strip()
        output_filename = gconfig.get_fixed("output_filename") or "generated_from_cases.can"
        cin_output_filename = (
            gconfig.get_fixed("cin_output_filename") or "generated_from_keyword.cin"
        )
        sheet_names_str = gconfig.get_fixed("mapping_sheets") or gconfig.get(
            "PATHS",
            "Mapping_Sheets",
            "HIL用例关键字说明,EM_CAN&Uart&LIN,EM_SOA,EM_总线测试专用,EM_设备&其他",
        )
        sheet_names = [s.strip() for s in (sheet_names_str or "").split(",") if s.strip()]

        testmode_dir = resolve_target_subdir(base_dir, output_dir, "TESTmode")
        testcases_dir = os.path.join(testmode_dir, "Testcases")
        os.makedirs(testcases_dir, exist_ok=True)
        master_output_path = os.path.join(testmode_dir, output_filename)

        # 是否包含 .cin 由「是否选择了关键字集 Clib 配置表」决定，而非域类型
        cin_excel_path = CANEntrypointSupport.resolve_cin_excel_path(
            gconfig, base_dir, domain=domain
        )
        has_keyword_clib = bool((cin_excel_path or "").strip())

        secoc_qualifier = CANEntrypointSupport._resolve_uds_qualifier(
            gconfig, base_dir, domain, output_dir
        )

        return {
            "domain": domain,
            "excel_files": excel_files,
            "mapping_excel_path": mapping_excel_path,
            "sheet_names": sheet_names,
            "testcases_dir": testcases_dir,
            "master_output_path": master_output_path,
            "cin_output_filename": cin_output_filename,
            "has_keyword_clib": has_keyword_clib,
            "secoc_qualifier": secoc_qualifier,
        }

    @staticmethod
    def load_clib_names_from_excel(excel_path: str) -> set[str]:
        names: set[str] = set()
        if not excel_path or not os.path.exists(excel_path):
            return names
        try:
            wb = load_workbook(excel_path, data_only=True, read_only=False)
        except Exception:
            return names
        try:
            ws = wb["Clib_Matrix"] if "Clib_Matrix" in wb.sheetnames else wb[wb.sheetnames[0]]
            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            header_norm = [("" if x is None else str(x).strip()).lower() for x in header]
            name_idx = None
            for i, h in enumerate(header_norm):
                if h == "name" or "name" in h:
                    name_idx = i
                    break
            if name_idx is None:
                name_idx = 1 if len(header_norm) > 1 else 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if name_idx < len(row):
                    val = ("" if row[name_idx] is None else str(row[name_idx]).strip()).lower()
                    if val:
                        names.add(val)
        finally:
            wb.close()
        return names

    @staticmethod
    def resolve_base_dir(base_dir: str | None = None) -> str:
        """解析工程根目录，统一使用 infra.filesystem.pathing.get_project_root。"""
        if base_dir is not None:
            return os.path.abspath(base_dir)
        return get_project_root(__file__)

    @staticmethod
    def resolve_config_path(base_dir: str, config_path: str | None = None) -> str:
        if config_path is not None:
            return config_path
        # 新目录结构：统一使用 base_dir/config/Configuration.txt
        default_path = os.path.join(base_dir, "config", "Configuration.txt")
        return default_path

    @staticmethod
    def load_generator_config(
        base_dir: str,
        config_path: str | None = None,
    ) -> GeneratorConfig:
        resolved_config_path = CANEntrypointSupport.resolve_config_path(base_dir, config_path)
        return GeneratorConfig(base_dir, config_path=resolved_config_path).load()

    @staticmethod
    def resolve_cin_excel_path(
        gconfig: GeneratorConfig,
        base_dir: str,
        *,
        domain: str = "LR_REAR",
    ) -> str:
        cin_excel_path = gconfig.get_first(
            CANEntrypointSupport._domain_candidates(
                domain,
                "cin_input_excel",
                "Cin_Input_Excel",
            )
        )
        if cin_excel_path and not os.path.isabs(cin_excel_path):
            cin_excel_path = os.path.abspath(os.path.join(base_dir, cin_excel_path))
        return cin_excel_path
