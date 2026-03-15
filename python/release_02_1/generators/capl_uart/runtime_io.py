#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
UART 生成运行期 IO：从 generate_uart_from_config 迁入的路径/配置/Excel/生成与日志实现。
供 runtime 导出，供 UARTGeneratorService 通过 _rt.xxx 调用。
"""

from __future__ import annotations

import configparser
import logging
import os
import re
import sys

from openpyxl import load_workbook

from core.generator_logging import GeneratorLogger
from infra.config import read_fixed_config
from core.parse_table_loggers import get_uart_matrix_logger
from utils.logger import TeeToLogger
from infra.filesystem.pathing import get_project_root
from utils.path_utils import find_config_path, resolve_target_subdir

# ---------- 模块级变量：解析表格日志，供 read_uart_excel_data 与 build_stdout_tee 使用 ----------
_parse_uart_logger: logging.Logger | None = None

# UART 通信矩阵必填列：内部名 -> 显示名，用于报错提示
_UART_REQUIRED_COLS_DISPLAY = {
    "msg_id": "Msg ID (Hex)",
    "message_name": "Message Name / 消息名称",
    "dlc": "DLC (Byte)",
    "signal_name": "Signal Name / 信号名称",
    "array": "Array",
    "length": "Length (Bits)",
    "value_type": "Value Type",
}


def _normalize_header(cell_value: str | None) -> str:
    """将表头单元格规范化：去掉换行、回车、制表符及首尾空白，连续空白压成单个空格，便于兼容误输入。"""
    if cell_value is None:
        return ""
    s = str(cell_value).strip()
    # 将任意空白字符（\n \r \t 及连续空格）统一为单个空格
    s = re.sub(r"[\s\u00a0]+", " ", s)
    return s.strip()


def flush_std_streams() -> None:
    """在 UART 运行前刷新 stdout/stderr，减少打包环境下的残留缓冲。"""
    if sys.stdout is not None:
        try:
            sys.stdout.flush()
        except (AttributeError, OSError):
            pass
    if sys.stderr is not None:
        try:
            sys.stderr.flush()
        except (AttributeError, OSError):
            pass


def resolve_runtime_paths() -> tuple[str, str]:
    """解析 UART 运行根目录与配置文件路径，统一使用 infra.filesystem.pathing.get_project_root。"""
    base_dir = get_project_root(__file__)
    config_path = find_config_path(base_dir)
    if not config_path or not os.path.exists(config_path):
        error_msg = f"未找到配置文件: {config_path}"
        print(error_msg, file=sys.stderr)
        flush_std_streams()
        raise FileNotFoundError(error_msg)
    return base_dir, config_path


_LOGGER: logging.Logger | None = None
_LOG_MANAGER: GeneratorLogger | None = None


def setup_logging(base_dir: str) -> logging.Logger:
    """
    初始化本脚本日志，写入 log/log_YYYYMMDD_HHMMSS/生成文件日志/generate_uart_from_config.log。
    """
    global _LOGGER, _LOG_MANAGER
    if _LOG_MANAGER is not None:
        _LOG_MANAGER.clear()
    _LOG_MANAGER = GeneratorLogger(
        base_dir,
        log_basename="generate_uart_from_config.log",
        logger_name="generate_uart_from_config",
    )
    logger = _LOG_MANAGER.setup()
    print(f"[uart] 日志已启用：{_LOG_MANAGER.primary_log_path}")
    _LOGGER = logger
    return logger


def get_parse_logger(base_dir: str) -> logging.Logger:
    """设置并返回 Uart_Matrix 解析表格 logger，供 build_stdout_tee 与 read_uart_excel_data 使用。"""
    global _parse_uart_logger
    _parse_uart_logger = get_uart_matrix_logger(base_dir)
    return _parse_uart_logger


def build_stdout_tee(logger: logging.Logger, stream):
    """将 stdout 绑定到 logger 与原 stream 的 Tee。"""
    global _parse_uart_logger
    if _parse_uart_logger is None:
        try:
            base_dir = get_project_root(__file__)
            _parse_uart_logger = get_uart_matrix_logger(base_dir)
        except Exception:
            pass
    return TeeToLogger(
        logger,
        logging.INFO,
        stream,
        error_prefixes=("[uart] 错误", "[错误]", "[error]"),
        warning_prefixes=("[uart] 警告", "[警告]", "[warn]"),
        start_trigger="FrameTypeIs8676",
        strip_whitespace=False,
    )


def load_config_with_repair(config_path: str, logger: logging.Logger) -> configparser.ConfigParser:
    """读取 UART 配置；若存在重复 option，则清理后重读。"""
    config = configparser.ConfigParser(allow_no_value=True)
    try:
        config.read(config_path, encoding="utf-8")
    except configparser.DuplicateOptionError:
        with open(config_path, "r", encoding="utf-8") as f:
            lines = f.readlines()
        seen_options = {}
        current_section = None
        cleaned_lines = []
        for line in lines:
            stripped = line.strip()
            if stripped.startswith("[") and stripped.endswith("]"):
                current_section = stripped[1:-1]
                seen_options[current_section] = set()
                cleaned_lines.append(line)
            elif "=" in stripped and not stripped.startswith("#"):
                key = stripped.split("=")[0].strip()
                if current_section and key:
                    if key.lower() in seen_options.get(current_section, set()):
                        continue
                    seen_options[current_section].add(key.lower())
                cleaned_lines.append(line)
            else:
                cleaned_lines.append(line)
        with open(config_path, "w", encoding="utf-8") as f:
            f.writelines(cleaned_lines)
        config.read(config_path, encoding="utf-8")
        if logger:
            logger.info("配置文件已修复，重复选项已移除")
    return config


def read_uart_rs232_config(config: configparser.ConfigParser) -> dict | None:
    """
    从配置的 [CENTRAL] 节读取串口通信参数，供 [UARTRS232] 节生成使用。
    只要存在任一 uart_comm_* 选项就返回 dict（空值用默认值填充），确保 Uart.txt 能生成 [UARTRS232] 段。
    无 CENTRAL 或没有任何 uart_comm_* 选项时返回 None。
    """
    if not config.has_section("CENTRAL"):
        return None

    has_any_config = (
        config.has_option("CENTRAL", "uart_comm_port")
        or config.has_option("CENTRAL", "uart_comm_baudrate")
        or config.has_option("CENTRAL", "uart_comm_dataBits")
        or config.has_option("CENTRAL", "uart_comm_stopBits")
        or config.has_option("CENTRAL", "uart_comm_kHANDSHAKE_DISABLED")
        or config.has_option("CENTRAL", "uart_comm_parity")
        or config.has_option("CENTRAL", "uart_comm_frameTypeIs8676")
    )
    if not has_any_config:
        return None

    # 有任一选项即构建配置（空值用默认值），保证 Uart.txt 一定写入 [UARTRS232]
    def _get(key: str, default: str = "") -> str:
        if not config.has_option("CENTRAL", key):
            return default
        v = config.get("CENTRAL", key, fallback="").strip()
        return v if v else default

    uart_rs232_config = {
        "port": _get("uart_comm_port", ""),
        "baudrate": _get("uart_comm_baudrate", "115200"),
        "dataBits": _get("uart_comm_dataBits", "8"),
        "stopBits": _get("uart_comm_stopBits", "1"),
        "kHANDSHAKE_DISABLED": _get("uart_comm_kHANDSHAKE_DISABLED", "0"),
        "parity": _get("uart_comm_parity", "0"),
        "frameTypeIs8676": _get("uart_comm_frameTypeIs8676", "0"),
    }
    return uart_rs232_config


def read_frame_type_value(config: configparser.ConfigParser) -> str:
    """从配置的 [PATHS] 节读取 FrameTypeIs8676 的值（0 或 1）。未配置或非 0/1 时返回 "0"。"""
    known_config_keys = {
        "Input_Excel", "Mapping_Excel", "Mapping_Sheets", "Output_FileName", "Cin_Output_FileName",
        "Cin_Input_Excel", "Cin_Input_Sheet", "Cin_Mapping_Excel", "Cin_Mapping_Sheet", "Output_Cin_FileName",
        "Xml_Input_Excel", "Xml_Output_FileName",
        "Uart_Input_Excel", "Uart_Output_FileName", "Output_Dir", "Output_Dir_Uart", "Output_Dir_Can", "Output_Dir_Cin", "Output_Dir_Xml",
    }
    frame_type_value = None
    if config.has_option("PATHS", "Uart_FrameTypeIs8676"):
        frame_type_value = config.get("PATHS", "Uart_FrameTypeIs8676")
    elif config.has_option("PATHS", "FrameTypeIs8676"):
        frame_type_value = config.get("PATHS", "FrameTypeIs8676")
    else:
        for key in config.options("PATHS"):
            if key in known_config_keys:
                continue
            value = config.get("PATHS", key).strip()
            if value in ["0", "1"]:
                frame_type_value = value
                break
    if frame_type_value is None:
        frame_type_value = "0"
    else:
        frame_type_value = str(frame_type_value).strip()
    if frame_type_value not in ["0", "1"]:
        frame_type_value = "0"
    return frame_type_value


def resolve_io_paths(config: configparser.ConfigParser, base_dir: str) -> tuple[str, str, str]:
    """从配置解析 UART 输入 Excel、输出文件名与输出目录。依赖 read_fixed_config(base_dir)。"""
    fixed = read_fixed_config(base_dir)
    input_excel = ""
    if config.has_section("CENTRAL"):
        input_excel = config.get("CENTRAL", "uart_excel", fallback="").strip()
    if not input_excel and config.has_section("PATHS"):
        input_excel = config.get(
            "PATHS",
            "Uart_Input_Excel",
            fallback="input/MCU_CDCU_CommunicationMatrix.xlsx",
        ).strip()
    if not input_excel:
        input_excel = "input/MCU_CDCU_CommunicationMatrix.xlsx"
    output_file = (fixed.get("uart_output_filename") or "Uart.txt").strip() or "Uart.txt"
    output_dir_uart = ""
    if config.has_section("CENTRAL"):
        output_dir_uart = config.get("CENTRAL", "output_dir", fallback="").strip()
    if not output_dir_uart and config.has_section("PATHS"):
        output_dir_uart = config.get("PATHS", "Output_Dir_Uart", fallback=None)
    if output_dir_uart and output_dir_uart.strip():
        output_dir = output_dir_uart.strip()
    else:
        output_dir = (
            config.get("PATHS", "Output_Dir", fallback="./output").strip()
            if config.has_section("PATHS")
            else "./output"
        )
    input_excel_normalized = input_excel.replace("/", os.sep)
    if not os.path.isabs(input_excel_normalized):
        input_excel = os.path.join(base_dir, input_excel_normalized)
    else:
        input_excel = input_excel_normalized
    input_excel = os.path.normpath(input_excel)
    output_dir = resolve_target_subdir(base_dir, output_dir, "Configuration")
    output_path = os.path.join(output_dir, output_file)
    return input_excel, output_file, output_path


def read_uart_excel_data(excel_path: str, sheet_name: str = "IVIToMCU") -> list:
    """
    从 UART 通信矩阵 Excel 指定 Sheet 读取消息与信号，按 Msg ID 聚合为消息列表。
    Sheet 不存在或缺少必填列时返回 []。
    """
    global _parse_uart_logger
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"找不到 Excel 文件: {excel_path}")

    try:
        wb = load_workbook(excel_path, data_only=True, read_only=True)
    except Exception as e:
        error_msg = str(e)
        if "decompressing" in error_msg.lower() or "incorrect header" in error_msg.lower() or "badzipfile" in error_msg.lower():
            raise ValueError(
                f"UART Excel 文件格式错误或文件已损坏: {excel_path}\n"
                f"错误详情: {error_msg}\n"
                f"请检查文件是否是有效的 Excel 文件（.xlsx 格式）"
            )
        raise ValueError(f"无法读取 UART Excel 文件: {excel_path}\n错误详情: {error_msg}")

    if sheet_name not in wb.sheetnames:
        err = f"Sheet '{sheet_name}' 不存在，可用的 Sheet: {wb.sheetnames}"
        print(f"[错误] UART 表头解析失败 - {sheet_name}: {err}")
        if _parse_uart_logger:
            _parse_uart_logger.error("UART 表头解析失败 - %s: %s", sheet_name, err)
        try:
            wb.close()
        except Exception:
            pass
        return []

    ws = wb[sheet_name]
    print(f"正在使用工作表: {sheet_name}")

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header = [str(h).strip() if h is not None else "" for h in header_row]

    col_indices = {
        "msg_id": None,
        "message_name": None,
        "dlc": None,
        "signal_name": None,
        "array": None,
        "length": None,
        "value_type": None,
    }
    for i, h in enumerate(header):
        # 规范化表头：兼容单元格内误输入的换行符、多余空格等
        h_norm = _normalize_header(h)
        h_lower = h_norm.lower()
        if "msg id" in h_lower and "hex" in h_lower:
            col_indices["msg_id"] = i
        elif "message name" in h_lower or "消息名称" in h_norm.replace(" ", ""):
            col_indices["message_name"] = i
        elif "dlc" in h_lower and "byte" in h_lower:
            col_indices["dlc"] = i
        elif "signal name" in h_lower or "信号名称" in h_norm.replace(" ", ""):
            col_indices["signal_name"] = i
        elif h_lower == "array":
            col_indices["array"] = i
        elif "length" in h_lower and "bits" in h_lower:
            col_indices["length"] = i
        elif "value type" in h_lower:
            col_indices["value_type"] = i

    required_cols = ["msg_id", "message_name", "dlc", "signal_name", "array", "length", "value_type"]
    missing_cols = [col for col in required_cols if col_indices[col] is None]
    if missing_cols:
        display_names = [_UART_REQUIRED_COLS_DISPLAY.get(c, c) for c in missing_cols]
        if _parse_uart_logger:
            _parse_uart_logger.error(
                "UART 通信矩阵表缺少必填列：excel=%s sheet=%s 缺少=%s",
                os.path.basename(excel_path),
                sheet_name,
                "、".join(display_names),
            )
        print(f"[错误] UART 表头解析失败 - {sheet_name}: 缺少必填列 {'、'.join(display_names)}，跳过该 sheet")
        print(f"  当前表头: {header}")
        print(f"  必填列要求: Msg ID (Hex)、Message Name / 消息名称、DLC (Byte)、Signal Name / 信号名称、Array、Length (Bits)、Value Type")
        try:
            wb.close()
        except Exception:
            pass
        return []

    messages = []
    current_msg = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        msg_id_value = row[col_indices["msg_id"]] if len(row) > col_indices["msg_id"] else None
        msg_id = str(msg_id_value).strip() if msg_id_value is not None else None
        message_name_value = row[col_indices["message_name"]] if len(row) > col_indices["message_name"] else None
        message_name = str(message_name_value).strip() if message_name_value is not None else None
        dlc_value = row[col_indices["dlc"]] if len(row) > col_indices["dlc"] else None
        dlc = str(dlc_value).strip() if dlc_value is not None else None
        signal_name_value = row[col_indices["signal_name"]] if len(row) > col_indices["signal_name"] else None
        signal_name = str(signal_name_value).strip() if signal_name_value is not None else None
        array_value = row[col_indices["array"]] if len(row) > col_indices["array"] else None
        array = str(array_value).strip() if array_value is not None else None
        length_value = row[col_indices["length"]] if len(row) > col_indices["length"] else None
        length = str(length_value).strip() if length_value is not None else None
        value_type_value = row[col_indices["value_type"]] if len(row) > col_indices["value_type"] else None
        value_type = str(value_type_value).strip() if value_type_value is not None else None

        if msg_id and msg_id != "None":
            if current_msg is not None:
                messages.append(current_msg)
            current_msg = {
                "msg_id": msg_id,
                "message_name": message_name if message_name and message_name != "None" else "",
                "dlc": dlc if dlc and dlc != "None" else "",
                "signals": [],
            }
        if signal_name and signal_name != "None" and current_msg is not None:
            signal = {
                "signal_name": signal_name,
                "array": array if array and array != "None" else "",
                "length": length if length and length != "None" else "",
                "value_type": value_type if value_type and value_type != "None" else "",
            }
            current_msg["signals"].append(signal)

    if current_msg is not None:
        messages.append(current_msg)
    return messages


def generate_uart_content(
    frame_type_value: str,
    ivi_to_mcu_messages: list,
    mcu_to_ivi_messages: list,
    uart_rs232_config: dict | None = None,
) -> str:
    """
    拼装 Uart.txt 全文：可选 [UARTRS232]、[IVIToMCU]、[MCUToIVI]。
    """
    lines = []

    if uart_rs232_config:
        # 上半部分仅生成 [UARTRS232]，格式：key=value//注释（与约定 uart.txt 格式一致）
        port = uart_rs232_config.get("port", "").strip()
        port_num = "0"
        if port:
            match = re.match(r"^COM(\d+)$", port, re.IGNORECASE)
            if match:
                port_num = match.group(1)
            elif port.isdigit():
                port_num = port
            else:
                digits = re.findall(r"\d+", port)
                port_num = digits[0] if digits else "0"

        baudrate = uart_rs232_config.get("baudrate") or "115200"
        dataBits = uart_rs232_config.get("dataBits") or "8"
        stopBits = uart_rs232_config.get("stopBits") or "1"
        handshake = uart_rs232_config.get("kHANDSHAKE_DISABLED")
        if handshake is None or handshake == "":
            handshake = "0"
        parity = uart_rs232_config.get("parity")
        if parity is None or parity == "":
            parity = "0"
        frame_type = uart_rs232_config.get("frameTypeIs8676") or frame_type_value or "0"

        lines.append("[UARTRS232]")
        lines.append(f"port={port_num}//端口号")
        lines.append(f"baudrate={baudrate}//波特率")
        lines.append(f"dataBits={dataBits}//数据位")
        lines.append(f"stopBits={stopBits}//停止位")
        lines.append(f"kHANDSHAKE_DISABLED={handshake}//握手")
        lines.append(f"parity={parity}//校验")
        lines.append(f"frameTypeIs8676={frame_type}")
        lines.append("")
        print(
            f" 已生成[UARTRS232]节: port={port_num}, baudrate={baudrate}, dataBits={dataBits}, stopBits={stopBits}, handshake={handshake}, parity={parity}, frameTypeIs8676={frame_type}",
            flush=True,
        )
    else:
        print("⚠ 警告: 未找到串口通信配置，跳过[UARTRS232]节生成", flush=True)
        print("   提示: 请在前端配置串口通信参数", flush=True)

    if ivi_to_mcu_messages:
        lines.append("[IVIToMCU]")
        for msg in ivi_to_mcu_messages:
            lines.append(f"Msg:0x{msg['msg_id']} {msg['message_name']} {msg['dlc']}")
            for signal in msg["signals"]:
                lines.append(f"{signal['signal_name']} {signal['array']} {signal['length']} {signal['value_type']}")
        lines.append("")

    if mcu_to_ivi_messages:
        lines.append("[MCUToIVI]")
        for msg in mcu_to_ivi_messages:
            lines.append(f"Msg:0x{msg['msg_id']} {msg['message_name']} {msg['dlc']}")
            for signal in msg["signals"]:
                lines.append(f"{signal['signal_name']} {signal['array']} {signal['length']} {signal['value_type']}")
        lines.append("")

    content = "\n".join(lines)
    content = re.sub(r"(\r?\n){2,}", r"\n", content)
    return content


def write_text_safe(output_path: str, content: str) -> None:
    """写入文本文件，utf-8 优先，失败时回退 gb18030。"""
    from utils import file_io

    file_io.write_text_safe(output_path, content, encoding="utf-8", fallback_encoding="gb18030")
