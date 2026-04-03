#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
交互服务层（GuiService）

目标：收纳 Tkinter 弹窗和文件/文件夹解析逻辑。
职责：封装 tk_lock，提供 select_path(type)、parse_file_structure。
"""

from __future__ import annotations

import gc
import os
import re
import threading
import traceback
import xml.etree.ElementTree as ET
from typing import Any, Dict, List, Literal, Optional

from openpyxl import load_workbook

try:
    import tkinter as tk
    from tkinter import filedialog
except ImportError:  # pragma: no cover - headless / minimal Python
    tk = None  # type: ignore[assignment]
    filedialog = None  # type: ignore[assignment]

# 线程锁：无 GUI 环境仍可通过 try/except 导入本模块
tk_lock = threading.Lock()


def _parse_excel_sheets(path: str) -> Dict[str, Any]:
    """解析 Excel 文件的 sheet 名称列表。
    参数：path — Excel 文件路径。
    返回：{"type": "excel", "sheets": [...]} 或 {"type": "excel", "error": str}。
    """
    if os.path.basename(path).startswith("~$"):
        return {"type": "excel", "sheets": []}
    try:
        wb = load_workbook(path, read_only=True)
        sheets = list(wb.sheetnames)
        wb.close()
        return {"type": "excel", "sheets": sheets}
    except Exception as error:
        return {"type": "excel", "error": str(error)}


def _parse_can_testcases(path: str) -> Dict[str, Any]:
    """解析 CAN 文件中的 testcase 名称列表。
    参数：path — .can 文件路径。
    返回：{"type": "can", "testcases": [...]} 或 {"type": "can", "error": str}。
    """
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            content = f.read()
        pattern = re.compile(r"testcase\s+(\w+)\s*\(", re.IGNORECASE)
        names = pattern.findall(content)
        return {"type": "can", "testcases": list(dict.fromkeys(names))}
    except Exception as error:
        return {"type": "can", "error": str(error)}


def _parse_xml_structure(path: str) -> Dict[str, Any]:
    """解析 XML 文件中的 testgroup 与 capltestcase 结构。
    参数：path — XML 文件路径。
    返回：{"type": "xml", "testgroups": [...], "capltestcases": [...]} 或 {"type": "xml", "error": str}。
    """
    try:
        tree = ET.parse(path)
        root = tree.getroot()
        result = {"type": "xml", "testgroups": [], "capltestcases": []}
        for elem in root.iter():
            if elem.tag.endswith("testgroup"):
                title = elem.get("title") or elem.get("ident") or ""
                if title:
                    result["testgroups"].append(title)
            elif elem.tag.endswith("capltestcase"):
                name = elem.get("name") or ""
                if name:
                    result["capltestcases"].append(name)
        return result
    except Exception as error:
        return {"type": "xml", "error": str(error)}


def _parse_file_structure_single(path: str) -> Dict[str, Any]:
    """解析单个文件的结构（Excel 的 sheet / CAN 的 testcase / XML 的 testgroup）。
    参数：path — 文件路径。
    返回：按类型返回 type + sheets|testcases|testgroups 或 error。
    """
    path_lower = path.lower()
    if path_lower.endswith((".xlsx", ".xlsm")):
        return _parse_excel_sheets(path)
    if path_lower.endswith(".can"):
        return _parse_can_testcases(path)
    if path_lower.endswith(".xml"):
        return _parse_xml_structure(path)
    return {"type": "unknown", "error": "不支持的文件格式，仅支持 Excel(.xlsx/.xlsm)、CAN(.can)、XML(.xml)"}


class GuiService:
    """
    GUI 交互服务：弹窗选路径、解析文件结构。
    使用 tk_lock 保证同一时间只有一个 Tk 弹窗，防止 Tcl 崩溃。
    """

    @staticmethod
    def config_filetypes() -> List[tuple[str, str]]:
        """配置文件选择器：仅 INI。"""
        return [
            ("INI 配置", "*.ini"),
            ("所有文件", "*.*"),
        ]

    @staticmethod
    def select_path(
        file_type: Literal["file", "folder"] = "file",
    ) -> Optional[str]:
        """弹出系统选择文件/文件夹窗口（线程安全 + 置顶）。
        参数：file_type — "file" 选文件（Excel），"folder" 选文件夹。
        返回：选中路径；取消或异常返回 None。
        """
        if tk is None or filedialog is None:
            print("当前环境无 tkinter，无法弹出选择框。")
            return None
        with tk_lock:
            root = None
            try:
                root = tk.Tk()
                root.withdraw()
                root.attributes("-topmost", True)
                root.update_idletasks()

                if file_type == "folder":
                    path = filedialog.askdirectory(parent=root, title="选择文件夹")
                else:
                    path = filedialog.askopenfilename(
                        parent=root,
                        title="选择 Excel 文件",
                        filetypes=[("Excel files", "*.xlsx;*.xlsm"), ("All files", "*.*")],
                    )
                return path or None
            except Exception as error:
                print(f"弹出选择框出错: {error}\n{traceback.format_exc()}")
                return None
            finally:
                if root:
                    try:
                        root.quit()
                    except Exception:
                        pass
                    try:
                        root.destroy()
                    except Exception:
                        pass
                gc.collect()

    @staticmethod
    def ask_saveas_filename(
        title: str = "保存配置文件",
        defaultextension: str = ".ini",
        initialfile: Optional[str] = None,
    ) -> Optional[str]:
        """弹出“另存为”对话框，返回用户选择的保存路径。
        参数：title — 窗口标题；defaultextension — 默认扩展名；initialfile — 初始文件名。
        返回：选中路径；取消或异常返回 None。
        """
        if tk is None or filedialog is None:
            print("当前环境无 tkinter，无法弹出另存为对话框。")
            return None
        with tk_lock:
            root = None
            try:
                root = tk.Tk()
                root.withdraw()
                root.attributes("-topmost", True)
                root.update_idletasks()
                path = filedialog.asksaveasfilename(
                    parent=root,
                    title=title,
                    defaultextension=defaultextension,
                    filetypes=GuiService.config_filetypes(),
                    initialfile=initialfile,
                )
                return path or None
            except Exception as error:
                print(f"另存为弹窗出错: {error}\n{traceback.format_exc()}")
                return None
            finally:
                if root:
                    try:
                        root.quit()
                    except Exception:
                        pass
                    try:
                        root.destroy()
                    except Exception:
                        pass
                gc.collect()

    @staticmethod
    def ask_open_config_filename(
        title: str = "选择要导入的配置文件",
        filetypes: Optional[List[tuple]] = None,
    ) -> Optional[str]:
        """弹出“打开文件”对话框，用于选择配置文件（.ini）。
        参数：title — 窗口标题；filetypes — 可选，文件类型列表，默认仅 .ini。
        返回：选中路径；取消或异常返回 None。
        """
        if tk is None or filedialog is None:
            print("当前环境无 tkinter，无法弹出打开文件对话框。")
            return None
        if filetypes is None:
            filetypes = GuiService.config_filetypes()
        with tk_lock:
            root = None
            try:
                root = tk.Tk()
                root.withdraw()
                root.attributes("-topmost", True)
                root.update_idletasks()
                path = filedialog.askopenfilename(
                    parent=root,
                    title=title,
                    filetypes=filetypes,
                )
                return path or None
            except Exception as error:
                print(f"打开文件弹窗出错: {error}\n{traceback.format_exc()}")
                return None
            finally:
                if root:
                    try:
                        root.quit()
                    except Exception:
                        pass
                    try:
                        root.destroy()
                    except Exception:
                        pass
                gc.collect()

    @staticmethod
    def parse_file_structure(path: str) -> Dict[str, Any]:
        """解析文件或文件夹下的 Excel/CAN/XML 结构（sheet、testcase、testgroup 等）。
        参数：path — 文件或文件夹路径；文件夹时遍历其下 Excel/CAN/XML。
        返回：{"success": True, "data": [...]} 或 {"success": False, "message": str}。
        """
        if not path or not path.strip():
            return {"success": False, "message": "未提供路径"}
        path = path.strip()
        if not os.path.exists(path):
            return {"success": False, "message": f"路径不存在: {path}"}

        results: List[Dict[str, Any]] = []
        try:
            if os.path.isfile(path):
                item = _parse_file_structure_single(path)
                item["filename"] = os.path.basename(path)
                results.append(item)
            else:
                for root_dir, _, files in os.walk(path):
                    for f in files:
                        if f.startswith("~$"):
                            continue
                        fp = os.path.join(root_dir, f)
                        fl = f.lower()
                        if fl.endswith((".xlsx", ".xlsm", ".can", ".xml")):
                            item = _parse_file_structure_single(fp)
                            item["filename"] = f
                            item["relpath"] = os.path.relpath(fp, path)
                            results.append(item)
            return {"success": True, "data": results}
        except Exception as error:
            return {"success": False, "message": str(error)}
