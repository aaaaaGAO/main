#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""DIDInfo 生成调度 service。不再依赖 DIDInfoLegacyHooks，通过 .runtime 委托根脚本实现。"""

from __future__ import annotations

import os
import sys
from pathlib import Path

import openpyxl

from . import runtime as _rt


class DIDInfoGeneratorService:
    """接管 DIDInfo 旧版主编排流程的 service。"""

    def run_legacy_pipeline(self):
        logger_obj = None
        base_dir = _rt.resolve_base_dir()
        logger_obj, old_stdout, old_stderr = _rt.init_runtime(base_dir)
        try:
            config_path, output_path, variant_names, inputs = _rt.load_runtime_config(base_dir)

            # 守卫 1：如果未配置任何输入 Excel，则直接跳过，不生成文件
            if not inputs:
                print("[didinfo] 未配置 ResetDid_Value 输入表，跳过 DIDInfo 生成。")
                return

            all_parts: list[str] = []
            last_sheet_name = None
            last_variant_name = None
            last_did = None
            last_len = None
            progress_level = _rt.get_progress_level()

            for excel_rel, sheets in inputs:
                excel_path_str = str(excel_rel).strip().replace("/", os.sep)
                excel_path = (
                    Path(excel_path_str)
                    if os.path.isabs(excel_path_str)
                    else (config_path.parent / Path(excel_path_str)).resolve()
                )
                if not excel_path.exists():
                    print(f"[didinfo] 警告: 未找到 Excel: {excel_path}，跳过")
                    continue

                try:
                    wb = openpyxl.load_workbook(excel_path, data_only=True, rich_text=True)
                except Exception as e:
                    print(f"[didinfo] 错误: 无法读取 Excel 文件: {excel_path}, 详情: {e}")
                    continue

                target_sheets = (
                    [_rt.pick_sheet_name(wb, None)]
                    if sheets is None
                    else (list(wb.sheetnames) if sheets == ["*"] else [s for s in sheets if s in wb.sheetnames])
                )

                for sheet_name in target_sheets:
                    ws = wb[sheet_name]
                    if logger_obj:
                        logger_obj.log(
                            progress_level,
                            f"解析 Excel 文件: {excel_path.name} sheet={sheet_name}",
                        )
                    try:
                        header_row, _ = _rt.find_header_row_and_cols(ws)
                    except RuntimeError:
                        print(
                            f"[didinfo] ERROR 跳过 sheet: Excel={excel_path.name} sheet={sheet_name} 表里少了必须列：Configure DID, Length (Bytes), Byte, Bit"
                        )
                        continue
                    try:
                        variant_cols = _rt.find_variant_cols(ws, header_row, variant_names)
                    except RuntimeError as e:
                        print(
                            f"[didinfo] ERROR 跳过 sheet: Excel={excel_path.name} sheet={sheet_name} 表里少了必须列（车型列）：{e}"
                        )
                        continue

                    sheet_parts = []
                    for vn in variant_names:
                        content, final_sheet, final_variant, final_did, final_len = _rt.generate_from_sheet(
                            ws,
                            excel_name=excel_path.name,
                            variant_name=vn,
                            variant_col=variant_cols[vn],
                            sheet_name=sheet_name,
                            last_sheet_name=last_sheet_name,
                            last_variant_name=last_variant_name,
                            last_did=last_did,
                            last_len=last_len,
                        )
                        sheet_parts.append(content.rstrip("\n"))
                        last_sheet_name, last_variant_name, last_did, last_len = (
                            final_sheet,
                            final_variant,
                            final_did,
                            final_len,
                        )

                    if sheet_parts:
                        all_parts.append("\n\n".join(sheet_parts))

            # 守卫 2：如果所有 Excel 中未解析出任何有效 DID 信息，则不生成文件
            if not all_parts:
                print("[didinfo] 未从输入的 Excel 中解析到有效 DID 信息，不生成文件。")
                return

            final_content = "\n\n".join([p for p in all_parts if p.strip()]) + "\n"
            output_path.parent.mkdir(parents=True, exist_ok=True)
            output_path.write_text(final_content, encoding="utf-8")
            print(f"已生成: {output_path}")
            if logger_obj:
                logger_obj.log(progress_level, f"已生成: {output_path}")
        finally:
            sys.stdout = old_stdout
            sys.stderr = old_stderr
            _rt.clear_run_logger(logger_obj)
