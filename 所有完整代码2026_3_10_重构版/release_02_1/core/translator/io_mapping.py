#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
IO Mapping 解析与替换（供 CAN/CIN 生成器复用）

功能概览：
  1. 从 Configuration.txt 的 [IOMAPPING] 段读取 Inputs，解析「Excel 路径 | Sheet 列表」。
  2. 打开一个或多个 IO_mapping Excel，支持指定 sheet 或全部 sheet（*）。
  3. 表头定位 Name/Path/Values；Name/Path/Values 均使用当前行单元格原始值，不做向下填充（不受合并单元格影响）。
  4. 将 Name -> Path 做替换；将 Values 中的「翻译前(右侧)」->「翻译后(左侧)」做枚举翻译（大小写不敏感）。
  5. 支持 J_DI*LS 特殊规则：二值枚举取反、分组分号、多参数透传等。

执行顺序（load_io_mapping_from_config 一次调用的步骤）：
  ① 读 config 的 [IOMAPPING].Inputs，解析为 (excel_path, sheets) 列表。
  ② 初始化日志（IO_Mapping.log），确定配置目录用于解析相对路径。
  ③ 对每个 Excel：校验路径与格式、打开工作簿；对每个 Sheet 定位表头（Name/Path/Values）。
  ④ 逐行读取 Name/Path/Values，填充 name_to_path、name_to_values；同表内冲突仅告警、保留首次。
  ⑤ 返回 IOMappingContext(name_to_path, name_to_values)，供上层对步骤参数调用 transform_args 做替换与枚举翻译。

代码阅读顺序（本文件已按此顺序组织，从上往下读即可）：
  - ① 常量与异常：PROGRESS_LEVEL、正则/集合常量、IOMappingParseError。
  - ② 读配置相关：get_base_dir、get_log_level_from_config、split_input_lines（来自 utils）。
  - ③ 日志相关：_IOProgressFormatter、_setup_logging、_log。
  - ④ 表头与列定位：_norm_header、_find_header_row_and_indices。
  - ⑤ 字符串/枚举工具：_find_colon、_is_numeric、_norm_key、_norm_enum_key、_has_expr_chars。
  - ⑥ Values 解析：_parse_values_cell。
  - ⑦ 上下文类：IOMappingContext（_maybe_invert_ls_enum、_is_j_di_ls、_process_inverted_token、transform_args）。
  - ⑧ 主加载入口：load_io_mapping_from_config。

配置建议（Configuration.txt）：
  [IOMAPPING]
  Inputs =
    input/xxx.xlsx | *
    input/yyy.xlsx | Sheet1,Sheet2
  注：不再检查 Enabled；只要配置了 Inputs 即启用。
"""

from __future__ import annotations

import logging
import logging.handlers
import os
import re
import unicodedata
from dataclasses import dataclass
from typing import Dict, List, Optional

from openpyxl import load_workbook

from core.caseid_log_dedup import DedupOnceFilter
from utils.excel_io import split_input_lines
from utils.logger import (
    ExcludeProgressFilter as _ExcludeProgressFilter_Shared,
    PROGRESS_LEVEL as _PROGRESS_LEVEL_SHARED,
    ProgressOnlyFilter as _ProgressOnlyFilter_Shared,
    get_log_level_from_config,
)
from utils.path_utils import get_base_dir

# ==================== ① 常量与异常 ====================

# 正常进度日志等级：不受 log_level_min 限制，始终写入（日志已启用、处理 Excel/sheet 等）
PROGRESS_LEVEL = _PROGRESS_LEVEL_SHARED

_ProgressOnlyFilter = _ProgressOnlyFilter_Shared
_ExcludeProgressFilter = _ExcludeProgressFilter_Shared

# 十进制数值、十六进制(0x/0X)、表达式符号、Values 冒号
_RE_NUMERIC = re.compile(r"^\s*[-+]?\d+(?:\.\d+)?\s*$")
_RE_HEX = re.compile(r"^\s*0[xX][0-9a-fA-F]+\s*$")
_EXPR_CHARS = set("><=()")
_COLON_CHARS = (":", "\uFF1A")  # 半角 : 、全角 ：

# J_DI*LS 多值枚举告警去重（同一 Name 只打一次）
_LS_INVERT_WARNED: set[str] = set()
_LOGGER: Optional[logging.Logger] = None


class IOMappingParseError(Exception):
    """当 IO mapping 替换/翻译失败时抛出，供上层生成注释行或报错。"""


# ==================== ② 读配置相关 ====================

# 使用 utils.path_utils.get_base_dir 替代 _default_base_dir
# 使用 utils.logger.get_log_level_from_config 替代 _get_log_level_from_config
# 使用 utils.excel_io.split_input_lines 替代 _split_mapping_input_lines


# ==================== ③ 日志相关 ====================


class _IOProgressFormatter(logging.Formatter):
    """进度类消息只输出时间+消息，不显示等级名。"""

    def format(self, record: logging.LogRecord) -> str:
        if record.levelno == PROGRESS_LEVEL:
            old_name = record.levelname
            record.levelname = " "
            s = super().format(record)
            record.levelname = old_name
            return s.replace("  ", " ", 1) if "  " in s else s
        return super().format(record)


def _setup_logging(base_dir: Optional[str]) -> logging.Logger:
    """
    初始化 IO_Mapping.log（写入 <base_dir>/log/，按大小轮转）。
    形参：base_dir - 传入工程根目录或 None（None 时用 get_base_dir()）。
    返回：logging.Logger。支持 log_level_min；进度类消息始终写入。
    """
    global _LOGGER
    base_dir = base_dir or get_base_dir()
    user_level = get_log_level_from_config(base_dir)

    from core.log_run_context import ensure_run_log_dirs

    run_dirs = ensure_run_log_dirs(base_dir)
    log_dir = run_dirs.parse_dir
    os.makedirs(log_dir, exist_ok=True)
    log_path = os.path.join(log_dir, "IO_Mapping.log")

    if _LOGGER is not None:
        desired = os.path.abspath(log_path)
        has_correct = any(
            isinstance(h, logging.FileHandler)
            and os.path.abspath(getattr(h, "baseFilename", "")) == desired
            for h in _LOGGER.handlers
        )
        if has_correct:
            return _LOGGER
        for h in _LOGGER.handlers[:]:
            try:
                h.close()
            except Exception:
                pass
            _LOGGER.removeHandler(h)

    logger = logging.getLogger("io_mapping")
    logger.setLevel(logging.DEBUG)
    logger.propagate = False

    if not logger.handlers:
        fmt = _IOProgressFormatter("%(asctime)s %(levelname)s %(message)s")
        dedup_filter = DedupOnceFilter()
        fh = logging.handlers.RotatingFileHandler(
            log_path,
            maxBytes=5 * 1024 * 1024,
            backupCount=20,
            encoding="utf-8",
        )
        fh.addFilter(_ExcludeProgressFilter())
        fh.addFilter(dedup_filter)
        fh.setLevel(user_level)
        fh.setFormatter(fmt)
        logger.addHandler(fh)
        fh_progress = logging.handlers.RotatingFileHandler(
            log_path,
            maxBytes=5 * 1024 * 1024,
            backupCount=20,
            encoding="utf-8",
        )
        fh_progress.addFilter(_ProgressOnlyFilter())
        fh_progress.addFilter(dedup_filter)
        fh_progress.setLevel(PROGRESS_LEVEL)
        fh_progress.setFormatter(fmt)
        logger.addHandler(fh_progress)
        ch = logging.StreamHandler()
        ch.setLevel(logging.INFO)
        ch.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
        ch.addFilter(dedup_filter)
        logger.addHandler(ch)

    _LOGGER = logger
    _LOGGER.log(PROGRESS_LEVEL, "[io_mapping] 日志已启用：%s", log_path)
    return logger


def _log(level: int, msg: str) -> None:
    """
    写日志或 fallback 到 print。
    形参：level - 传入 logging 等级（如 logging.ERROR）；msg - 传入日志消息字符串。
    返回：无。
    """
    if _LOGGER is not None:
        _LOGGER.log(level, msg)
    else:
        print(msg)


# ==================== ④ 表头与列定位 ====================


def _norm_header(v) -> str:
    """
    表头规范化：去首尾空白、移除空格、转小写，用于匹配 Name/Path/Values。
    形参：v - 传入表头单元格原始值（可为 None）。
    返回：str。空表头为 ""。
    """
    if v is None:
        return ""
    return str(v).strip().replace(" ", "").casefold()


def _find_header_row_and_indices(ws, *, max_scan_rows: int = 30) -> tuple[int, dict[str, int], list[str]]:
    """
    在工作表前若干行中定位同时包含 Name/Path/Values 的表头行。
    形参：ws - 传入 openpyxl 的 Worksheet 对象；max_scan_rows - 传入最大扫描行数，默认 30。
    返回：(header_row, col_map, missing)。
      - header_row: 表头行号（1-based），找不到为 -1。
      - col_map: {"name": idx, "path": idx, "values": idx}，idx 为 0-based 列索引。
      - missing: 缺失的必须列名列表（如 ["Name", "Path"]）。
    """
    required = {"name": "Name", "path": "Path", "values": "Values"}
    seen_headers: set[str] = set()
    max_r = min(getattr(ws, "max_row", 0) or 0, max_scan_rows) or max_scan_rows
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=max_r, values_only=True), start=1):
        found: dict[str, int] = {}
        for c0, cell_v in enumerate(row):
            key = _norm_header(cell_v)
            if not key or key not in required:
                continue
            seen_headers.add(key)
            if key not in found:
                found[key] = c0
        if all(k in found for k in required.keys()):
            return r, {"name": found["name"], "path": found["path"], "values": found["values"]}, []
    missing = [required[k] for k in required.keys() if k not in seen_headers]
    return -1, {}, missing


# ==================== ⑤ 字符串/枚举工具 ====================


def _find_colon(s: str, start: int) -> int:
    """
    返回 s[start:] 中第一个冒号（英文或中文）的位置。
    形参：s - 传入待查字符串；start - 传入起始下标。
    返回：int。位置索引，无则 -1。
    """
    candidates = [s.find(c, start) for c in _COLON_CHARS]
    candidates = [p for p in candidates if p >= 0]
    return min(candidates) if candidates else -1


def _is_numeric(s: str) -> bool:
    """
    判断字符串是否为数值（十进制或十六进制如 0x1）。
    形参：s - 传入待判断字符串（可为 None）。
    返回：bool。
    """
    if s is None:
        return False
    s_str = str(s).strip()
    return bool(_RE_NUMERIC.match(s_str) or _RE_HEX.match(s_str))


def _norm_key(s: str) -> str:
    """
    Name/Path 等键的规范化：去首尾空白、转小写，用于字典查找。
    形参：s - 传入原始键字符串。
    返回：str。
    """
    return str(s).strip().casefold()


def _norm_enum_key(s: str) -> str:
    """
    Values/枚举专用 key 规范化：去首尾空白、折叠内部空白、移除不可见字符、转小写。
    形参：s - 传入枚举显示值（如 "AUTO DOWN"）。
    返回：str。用于避免复制粘贴导致的匹配失败。
    """
    if s is None:
        return ""
    t = str(s).strip()
    t = "".join(ch for ch in t if unicodedata.category(ch) != "Cf")
    t = re.sub(r"\s+", " ", t)
    return t.casefold()


def _has_expr_chars(s: str) -> bool:
    """
    判断是否包含表达式符号 ><=()，此类字符串不做枚举翻译、直接透传。
    形参：s - 传入待判断字符串（可为 None）。
    返回：bool。
    """
    if s is None:
        return False
    return any((ch in _EXPR_CHARS) for ch in str(s))


# ==================== ⑥ Values 解析 ====================


def _parse_values_cell(values_cell: str) -> Dict[str, str]:
    """
    解析 Values 单元格，得到 翻译前(右侧) -> 翻译后(左侧) 的映射。
    支持多行、一行多对（如 1:ON 0:OFF）、数值 key 紧挨（如 0:AUTO DOWN2.2:AUTO UP）的强容错切分。
    形参：values_cell - 传入 Values 列单元格的原始字符串（可为 None）。
    返回：Dict[str, str]。key 为 _norm_enum_key(翻译前)，value 为翻译后（如 "1"、"0"）。
    """
    if values_cell is None:
        return {}
    text = str(values_cell).strip()
    if not text:
        return {}

    mapping: Dict[str, str] = {}
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue

        num_pat = re.compile(r"[-+]?\d+(?:\.\d+)?\s*[:\uFF1A]")
        ms = list(num_pat.finditer(line))
        if ms:
            for idx, m in enumerate(ms):
                colon_pos = m.end() - 1
                if idx == 0:
                    pair_start = 0
                else:
                    pos = m.start() - 1
                    while pos >= 0 and not line[pos].isalpha():
                        pos -= 1
                    if pos < 0:
                        pair_start = 0
                    else:
                        pos += 1
                        while pos < len(line) and line[pos].isspace():
                            pos += 1
                        pair_start = pos
                left = line[pair_start:colon_pos].strip()
                start = m.end()
                end = ms[idx + 1].start() if idx + 1 < len(ms) else len(line)
                right = line[start:end].strip()
                if not left or not right:
                    continue
                mapping[_norm_enum_key(right)] = left
            continue

        i = 0
        n = len(line)
        while i < n:
            j = _find_colon(line, i)
            if j < 0:
                break
            left = line[i:j].strip()
            if not left:
                i = j + 1
                continue
            k = j + 1
            while k < n and line[k].isspace():
                k += 1
            start_right = k
            m = re.search(r"\s+\S+\s*[:\uFF1A]", line[start_right:])
            next_pair_pos = start_right + m.start() if m else None
            if next_pair_pos is None:
                right = line[start_right:].strip()
                i = n
            else:
                right = line[start_right:next_pair_pos].strip()
                i = next_pair_pos
                while i < n and line[i].isspace():
                    i += 1
            if not right:
                continue
            mapping[_norm_enum_key(right)] = left.strip()
    return mapping


# ==================== ⑦ 上下文类 ====================


@dataclass
class IOMappingContext:
    """
    加载后的 IO mapping 上下文，供 CAN/CIN 步骤里对参数做 Name->Path 替换与 Values 枚举翻译。
    属性：name_to_path - Name(规范键)->Path；name_to_values - Name(规范键)->{枚举key->翻译后值}。
    """

    name_to_path: Dict[str, str]
    name_to_values: Dict[str, Dict[str, str]]

    @staticmethod
    def _maybe_invert_ls_enum(name: str, value_token: str, values_map: Dict[str, str]) -> str:
        """
        J_DI*LS 二值枚举取反：若 Name 为 J_DI*LS 且 Values 仅 2 个枚举，则返回「另一个」枚举 key。
        形参：name - 传入当前 Name（如 J_DI_RDoorInnerSw_LS）；value_token - 传入用户写的枚举值；values_map - 传入该 Name 的 Values 映射（norm_enum_key -> 翻译后）。
        返回：str。若无需取反或非二值，返回原 value_token；否则返回另一个枚举的 key。
        """
        if not name or not value_token or not values_map:
            return value_token
        n = str(name).strip().upper()
        if not (n.startswith("J_DI") and n.endswith("LS")):
            return value_token
        v = str(value_token).strip()
        if not v:
            return value_token
        keys = list(values_map.keys())
        cur = _norm_enum_key(v)
        if len(keys) != 2:
            if len(keys) > 2 and cur in values_map:
                if n not in _LS_INVERT_WARNED:
                    _LS_INVERT_WARNED.add(n)
                    enum_preview = ", ".join(sorted(keys)[:20])
                    more = "" if len(keys) <= 20 else f" ...(+{len(keys)-20})"
                    _log(
                        logging.ERROR,
                        "[io_mapping][ERROR] J_DI*LS 枚举反转仅支持 2 值，但该 Name 的 Values 有 "
                        f"{len(keys)} 值：name={name!r} enums=[{enum_preview}{more}] 当前输入={value_token!r}（本次将不反转）",
                    )
            return value_token
        if cur not in values_map:
            return value_token
        return keys[0] if keys[1] == cur else keys[1]

    @staticmethod
    def _is_j_di_ls(name: str) -> bool:
        """
        判断 Name 是否以 J_DI 开头且以 LS 结尾。
        形参：name - 传入 Name 字符串。
        返回：bool。
        """
        if not name:
            return False
        n = str(name).strip().upper()
        return n.startswith("J_DI") and n.endswith("LS")

    def _process_inverted_token(
        self,
        raw_name: str,
        value_token: str,
        values_map: Dict[str, str],
    ) -> str:
        """
        统一处理 J_DI*LS 的取反：数字 0/1 互换；枚举经 _maybe_invert_ls_enum 后按 values_map 翻译；表达式透传。
        形参：raw_name - 传入原始 Name；value_token - 传入当前 token；values_map - 传入该 Name 的 Values 映射。
        返回：str。翻译后或取反后的值。失败抛 IOMappingParseError。
        """
        if value_token is None:
            return value_token
        token = str(value_token).strip()
        if not token:
            return token
        if _has_expr_chars(token):
            return token
        if _is_numeric(token):
            try:
                num_val = int(token, 16) if _RE_HEX.match(token) else int(float(token))
            except (ValueError, OverflowError):
                return token
            if num_val == 0:
                return "1"
            if num_val == 1:
                return "0"
            return token
        if not values_map:
            raise IOMappingParseError(f"Values 为空，无法翻译: {token}")
        phrase2 = self._maybe_invert_ls_enum(raw_name, token, values_map)
        k = _norm_enum_key(phrase2)
        if k in values_map:
            return values_map[k]
        raise IOMappingParseError(
            f"J_DI*LS 名称 {raw_name!r} 的值必须是 0、1 或 Values 中的枚举值，"
            f"但收到 {token!r}。Values 中的有效枚举值: {list(values_map.keys())}"
        )

    def transform_args(self, args: List[str]) -> List[str]:
        """
        对关键字后的 args 做 IO mapping：args[0] 替换为 Path；第一个 value 做 Values 翻译（支持多词枚举、J_DI*LS 取反与分组）。
        形参：args - 传入步骤参数列表，args[0] 为以 J_ 开头的 Name，后续为 value 及可选单位等。
        返回：List[str]。首元素为 Path，其余为翻译后 value + 透传的后续 token。失败抛 IOMappingParseError。
        """
        if not args:
            raise IOMappingParseError("参数为空")
        raw_name = str(args[0]).strip()
        if not raw_name.upper().startswith("J_"):
            return args
        name_key = _norm_key(raw_name)
        if name_key not in self.name_to_path and name_key not in self.name_to_values:
            raise IOMappingParseError(f"Name 未找到: {raw_name}")
        path = self.name_to_path.get(name_key, "")
        if not path:
            raise IOMappingParseError(f"Path 为空: {raw_name}")
        values_map = self.name_to_values.get(name_key, {})
        out: List[str] = [path]
        if len(args) == 1:
            return out
        rest_tokens: List[str] = [str(raw).strip() for raw in args[1:] if str(raw).strip()]
        if not rest_tokens:
            return out
        if not values_map:
            first = rest_tokens[0]
            if _is_numeric(first) or _has_expr_chars(first):
                out.extend(rest_tokens)
                return out
            raise IOMappingParseError(f"Values 为空: Name={raw_name}, Value={' '.join(rest_tokens)}")

        if self._is_j_di_ls(raw_name):
            rest_str = " ".join(rest_tokens).strip()
            if not rest_str:
                return out
            group_strs = re.split(r"[;；]", rest_str)
            new_groups: List[str] = []
            for g in group_strs:
                g = g.strip()
                if not g:
                    continue
                seg_tokens = g.split()
                if len(seg_tokens) == 1:
                    new_groups.append(self._process_inverted_token(raw_name, seg_tokens[0], values_map))
                else:
                    inv = self._process_inverted_token(raw_name, seg_tokens[1], values_map)
                    new_groups.append(" ".join([seg_tokens[0], inv] + seg_tokens[2:]))
            new_rest_str = ";".join(new_groups)
            out.extend([tok for tok in new_rest_str.split() if tok])
            return out

        first = rest_tokens[0]
        if _has_expr_chars(first):
            out.extend(rest_tokens)
            return out
        if _is_numeric(first):
            if self._is_j_di_ls(raw_name):
                try:
                    num_val = (
                        int(str(first).strip(), 16)
                        if _RE_HEX.match(str(first).strip())
                        else int(float(first))
                    )
                    if num_val == 0:
                        out.append("1")
                        out.extend(rest_tokens[1:])
                        return out
                    if num_val == 1:
                        out.append("0")
                        out.extend(rest_tokens[1:])
                        return out
                    raise IOMappingParseError(
                        f"J_DI*LS 名称 {raw_name!r} 的值必须是 0、1 或 Values 中的枚举值，但收到数字 {num_val!r}。"
                    )
                except (ValueError, OverflowError):
                    pass
            out.extend(rest_tokens)
            return out

        max_n = 0
        for t in rest_tokens:
            if not t:
                continue
            if _is_numeric(t) or _has_expr_chars(t):
                break
            max_n += 1
        if max_n <= 0:
            out.extend(rest_tokens)
            return out
        matched = False
        for n in range(max_n, 0, -1):
            phrase = " ".join(rest_tokens[:n]).strip()
            if not phrase:
                continue
            phrase2 = self._maybe_invert_ls_enum(raw_name, phrase, values_map)
            k = _norm_enum_key(phrase2)
            if k in values_map:
                out.append(values_map[k])
                out.extend(rest_tokens[n:])
                matched = True
                break
        if not matched:
            if self._is_j_di_ls(raw_name):
                raise IOMappingParseError(
                    f"J_DI*LS 名称 {raw_name!r} 的值必须是 0、1 或 Values 中的枚举值，但收到 {first!r}。有效枚举: {list(values_map.keys())}"
                )
            raise IOMappingParseError(f"Values 未匹配: {first}")
        return out


# ==================== ⑧ 主加载入口 ====================


def _get_io_mapping_inputs_text(config, domain: str) -> str:
    """从配置中按域读取 [IOMAPPING] 的 Inputs 文本。参数: config — 配置对象；domain — 域（LR_REAR 兼容全局 IOMAPPING）。返回: Inputs 字符串。"""
    section_candidates = [f"{domain}_IOMAPPING"] if domain and domain != "LR_REAR" else ["IOMAPPING"]
    for section in section_candidates:
        if not config.has_section(section):
            continue
        inputs_text = config.get(section, "Inputs", fallback="") or config.get(
            section, "inputs", fallback=""
        )
        if inputs_text and str(inputs_text).strip():
            return str(inputs_text)
    return ""


def load_io_mapping_from_config(
    config,
    base_dir: Optional[str] = None,
    config_path: Optional[str] = None,
    domain: str = "LR_REAR",
) -> Optional[IOMappingContext]:
    """
    从 config 的 [IOMAPPING] 段加载 IO mapping；未配置 Inputs 则返回 None。
    形参：
      config - 传入 ConfigParser 对象（已读入 Configuration.txt）。
      base_dir - 传入工程根目录，用于日志与相对路径；可选。
      config_path - 传入配置文件路径，用于解析 Inputs 中相对路径的基准目录；可选，优先于 base_dir。
    返回：IOMappingContext 或 None。有 Inputs 时返回上下文，供 transform_args 使用。
    """
    inputs_text = _get_io_mapping_inputs_text(config, domain)
    input_lines = split_input_lines(inputs_text)
    if not input_lines:
        return None

    logger_base = base_dir or (
        os.path.dirname(os.path.abspath(config_path)) if config_path else None
    )
    logger = _setup_logging(logger_base)
    config_dir = (
        os.path.dirname(os.path.abspath(config_path))
        if config_path
        else (base_dir or os.getcwd())
    )

    name_to_path: Dict[str, str] = {}
    name_to_values: Dict[str, Dict[str, str]] = {}

    for excel_path_raw, sheets_raw in input_lines:
        excel_path = excel_path_raw.strip()
        excel_path_for_check = excel_path.replace("/", os.sep)
        if not os.path.isabs(excel_path_for_check):
            excel_path_for_check = os.path.abspath(
                os.path.join(config_dir, excel_path_for_check)
            )
        excel_path_for_check = os.path.normpath(excel_path_for_check)
        if not os.path.exists(excel_path_for_check):
            try:
                excel_path_utf8 = excel_path_for_check.encode("utf-8").decode("utf-8")
                if os.path.exists(excel_path_utf8):
                    excel_path_for_check = excel_path_utf8
                else:
                    raise FileNotFoundError(
                        f"找不到 IO_mapping Excel 文件: {excel_path_for_check}"
                    )
            except Exception:
                raise FileNotFoundError(f"找不到 IO_mapping Excel 文件: {excel_path_for_check}")
        if not os.path.isfile(excel_path_for_check):
            raise IOMappingParseError(f"路径不是文件: {excel_path_for_check}")
        if not excel_path_for_check.lower().endswith((".xlsx", ".xlsm")):
            raise IOMappingParseError(
                f"文件不是有效的 Excel 文件（.xlsx 或 .xlsm）: {excel_path_for_check}"
            )

        excel_path_to_open = excel_path_for_check
        try:
            with open(excel_path_to_open, "rb") as f:
                if f.read(4) != b"PK\x03\x04":
                    raise IOMappingParseError(
                        f"文件不是有效的 Excel（缺少 ZIP 文件头）: {excel_path_to_open}"
                    )
            wb = load_workbook(excel_path_to_open, data_only=True, read_only=False)
        except FileNotFoundError:
            raise IOMappingParseError(f"找不到文件: {excel_path_to_open}")
        except PermissionError:
            raise IOMappingParseError(f"没有权限读取: {excel_path_to_open}")
        except Exception as e:
            err = str(e).lower()
            if (
                "decompressing" in err
                or "incorrect header" in err
                or "badzipfile" in err
                or "not a zip file" in err
            ):
                raise IOMappingParseError(
                    f"IO Mapping Excel 格式错误或已损坏: {excel_path_to_open}\n{e}"
                )
            raise IOMappingParseError(f"无法读取 IO Mapping Excel: {excel_path_to_open}\n{e}")

        sheet_names = (
            [s for s in wb.sheetnames]
            if (sheets_raw.strip() == "*" or not sheets_raw.strip())
            else [s.strip() for s in sheets_raw.split(",") if s.strip()]
        )
        for sheet_name in sheet_names:
            if sheet_name not in wb.sheetnames:
                raise FileNotFoundError(
                    f"IO_mapping sheet 不存在: {excel_path_to_open} | {sheet_name}"
                )
            ws = wb[sheet_name]
            excel_name_for_log = os.path.basename(excel_path_to_open)
            header_row, col_map, missing = _find_header_row_and_indices(
                ws, max_scan_rows=30
            )
            if header_row < 0 or missing:
                logger.error(
                    "[io_mapping] 跳过sheet: Excel=%s sheet=%s 表头缺少必须列: %s",
                    excel_name_for_log,
                    sheet_name,
                    ", ".join(missing) if missing else "Name, Path, Values",
                )
                continue
            logger.log(
                PROGRESS_LEVEL,
                "[io_mapping] 处理Excel=%s sheet名=%s",
                excel_name_for_log,
                sheet_name,
            )
            idx_name, idx_path, idx_values = (
                col_map["name"],
                col_map["path"],
                col_map["values"],
            )
            local_name_to_path: Dict[str, str] = {}
            local_name_to_values: Dict[str, Dict[str, str]] = {}
            multi_path_warn: Dict[str, dict] = {}
            for excel_row, row in enumerate(
                ws.iter_rows(min_row=header_row + 1, values_only=True),
                start=header_row + 1,
            ):
                name = row[idx_name] if idx_name < len(row) else None
                path = row[idx_path] if idx_path < len(row) else None
                values_cell = row[idx_values] if idx_values < len(row) else None
                name_s = str(name).strip() if name is not None else ""
                path_s = str(path).strip() if path is not None else ""
                values_s = str(values_cell).strip() if values_cell is not None else ""
                if not name_s and not path_s and not values_s:
                    continue
                if not name_s:
                    continue
                k = _norm_key(name_s)
                src = f"{os.path.basename(excel_path)}/{sheet_name}"
                if path_s:
                    if k in local_name_to_path and local_name_to_path[k] != path_s:
                        rec = multi_path_warn.get(k)
                        if rec is None:
                            rec = {
                                "name": name_s,
                                "path_first": local_name_to_path.get(k, ""),
                                "ignored_paths": set(),
                                "rows": [],
                                "src": src,
                            }
                            multi_path_warn[k] = rec
                        rec["ignored_paths"].add(path_s)
                        rec["rows"].append(excel_row)
                    else:
                        local_name_to_path[k] = path_s
                        if k not in name_to_path:
                            name_to_path[k] = path_s
                if values_s:
                    vm = _parse_values_cell(values_s)
                    if vm:
                        cur_local = local_name_to_values.get(k)
                        if cur_local is None:
                            cur_local = {}
                            local_name_to_values[k] = cur_local
                        conflict_k = None
                        for enum_key, enum_val in vm.items():
                            if enum_key in cur_local and cur_local[enum_key] != enum_val:
                                conflict_k = (enum_key, cur_local[enum_key], enum_val)
                                break
                        if conflict_k is not None:
                            ek, v_old, v_new = conflict_k
                            logger.warning(
                                "[io_mapping] Name 同表 Values 冲突，保留首次：来源=%s 行=%s name=%r enum=%r val_first=%r val_ignored=%r",
                                src,
                                excel_row,
                                name_s,
                                ek,
                                v_old,
                                v_new,
                            )
                        else:
                            for enum_key, enum_val in vm.items():
                                if enum_key not in cur_local:
                                    cur_local[enum_key] = enum_val
                            cur_global = name_to_values.get(k)
                            if cur_global is None:
                                name_to_values[k] = dict(vm)
                            else:
                                for enum_key, enum_val in vm.items():
                                    if enum_key not in cur_global:
                                        cur_global[enum_key] = enum_val
            for _, rec in multi_path_warn.items():
                ignored_paths = sorted(rec["ignored_paths"])
                rows = sorted(set(rec["rows"]))
                rows_s = ",".join(str(x) for x in rows[:100]) + (
                    "" if len(rows) <= 100 else f"...(+{len(rows)-100})"
                )
                logger.warning(
                    "[io_mapping] Name 同表多 Path，保留首次：来源=%s name=%r path_first=%r path_ignored=%s 行=%s",
                    rec.get("src", ""),
                    rec.get("name", ""),
                    rec.get("path_first", ""),
                    ignored_paths,
                    rows_s,
                )

    return IOMappingContext(name_to_path=name_to_path, name_to_values=name_to_values)

